#!/usr/bin/env python3
"""
Compare Block 3 actuals (from Amion HTML schedules) against prior Block 1&2
actuals and annual targets from the Google Sheet.

Parses monthly HTML schedules for March–June 2026, classifies services using
the same rules as recalculate_prior_actuals.py, and produces a spreadsheet
with annual targets, Block 1&2 actuals, and Block 3 actuals side by side.

Output:
  output/block3_actuals.xlsx  — single sheet for Google Sheet paste

Usage:
    python -m analysis.compare_block3_actuals
"""

import csv
import io
import json
import os
import re
import sys
import urllib.request
from collections import defaultdict
from datetime import date

# Ensure project root is on path
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(SCRIPT_DIR)
sys.path.insert(0, PROJECT_ROOT)

from parse_schedule import parse_schedule, merge_schedules
from name_match import to_canonical, match_provider, clean_html_provider
from block.recalculate_prior_actuals import classify_service, parse_date

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

BLOCK_3_START = date(2026, 3, 2)   # Monday
BLOCK_3_END = date(2026, 6, 28)    # Sunday

MONTHLY_SCHEDULES_DIR = os.path.join(PROJECT_ROOT, "input", "monthlySchedules")
OUTPUT_DIR = os.path.join(PROJECT_ROOT, "output")

BLOCK_3_FILES = [
    "2026-03.html",
    "2026-04.html",
    "2026-05.html",
    "2026-06.html",
]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def in_block3_range(d):
    """Check if date falls within Block 3 (March 2 – June 28, 2026)."""
    if d is None:
        return False
    return BLOCK_3_START <= d <= BLOCK_3_END


def parse_float(val, default=0.0):
    """Parse a float value from CSV, returning default if empty/invalid."""
    if val is None:
        return default
    val = str(val).strip()
    if not val:
        return default
    try:
        return float(val)
    except ValueError:
        return default


# ---------------------------------------------------------------------------
# Google Sheet / CSV provider data
# ---------------------------------------------------------------------------

SHEET_ID = "1dbHUkE-pLtQJK02ig2EbU2eny33N60GQUvk4muiXW5M"
SHEET_BASE = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/gviz/tq?tqx=out:csv"


def load_providers_sheet():
    """Load provider data from Google Sheet (or local CSV fallback).

    Returns:
        (provider_order, provider_data) where:
        - provider_order is a list of provider names in Sheet row order
        - provider_data is a dict keyed by provider_name with annual targets
    """
    text = None

    # Try Google Sheet first
    try:
        url = f"{SHEET_BASE}&sheet={urllib.request.quote('Providers')}"
        print("  Fetching provider list from Google Sheet...")
        with urllib.request.urlopen(url, timeout=10) as resp:
            text = resp.read().decode("utf-8")
        # Cache locally
        cache_path = os.path.join(OUTPUT_DIR, "providers_sheet.csv")
        with open(cache_path, "w", newline="") as f:
            f.write(text)
        print("  Fetched from Google Sheet (cached to providers_sheet.csv)")
    except Exception as e:
        print(f"  WARNING: Could not fetch Google Sheet: {e}")

    # Fall back to local CSV
    if text is None:
        for fname in ("providers_sheet_updated.csv", "providers_sheet.csv"):
            csv_path = os.path.join(OUTPUT_DIR, fname)
            if os.path.isfile(csv_path):
                with open(csv_path, newline="") as f:
                    text = f.read()
                print(f"  Using local fallback: {fname}")
                break

    if text is None:
        print("  ERROR: No providers sheet available!")
        return [], {}

    reader = csv.DictReader(io.StringIO(text))
    provider_order = []
    provider_data = {}
    for row in reader:
        name = row.get("provider_name", "").strip()
        if not name:
            continue
        provider_order.append(name)
        provider_data[name] = {
            "annual_weeks": parse_float(row.get("annual_weeks")),
            "annual_weekends": parse_float(row.get("annual_weekends")),
            "annual_nights": parse_float(row.get("annual_nights")),
        }

    print(f"  Loaded {len(provider_order)} providers")
    return provider_order, provider_data


# ---------------------------------------------------------------------------
# Block 3 actuals computation
# ---------------------------------------------------------------------------

def compute_block3_actuals():
    """Parse Block 3 HTML schedules and count shifts per provider.

    Returns dict keyed by canonical provider name:
        {
            "weekday_shifts": int,
            "block3_weeks": float,
            "weekend_shifts": int,
            "block3_weekends": float,
            "night_shifts": int,
            "block3_nights": int,
            "swing_shifts": int,
        }
    """
    # Parse HTML files
    all_month_data = []
    for fname in BLOCK_3_FILES:
        fpath = os.path.join(MONTHLY_SCHEDULES_DIR, fname)
        if not os.path.exists(fpath):
            print(f"  WARNING: Missing file: {fname}")
            continue
        print(f"  Parsing: {fname}")
        month_data = parse_schedule(fpath)
        all_month_data.append(month_data)

    if not all_month_data:
        print("ERROR: No Block 3 schedule files found!")
        sys.exit(1)

    # Merge all months
    merged = merge_schedules(all_month_data)
    print(f"\nParsed {len(merged['schedule'])} days, "
          f"{len(merged.get('by_provider', {}))} unique providers")

    # Classify all services
    service_classes = {}
    for s in merged["services"]:
        cls = classify_service(s["name"], s["hours"])
        service_classes[s["name"]] = cls

    # Report classification counts
    counts_by_class = defaultdict(int)
    for cls in service_classes.values():
        counts_by_class[cls] += 1
    print(f"\nService classification: {counts_by_class['day']} day, "
          f"{counts_by_class['night']} night, {counts_by_class['swing']} swing, "
          f"{counts_by_class['exclude']} excluded")

    # Count shifts per provider, DEDUPLICATED per day
    # Priority: night(3) > swing(2) > weekday/weekend(1)
    provider_day_shifts = defaultdict(dict)  # provider -> {date_str: shift_type}
    dates_outside_range = 0
    moonlighting_excluded = 0

    for day_entry in merged["schedule"]:
        dow = day_entry["day_of_week"]
        is_weekend = dow in ("Sat", "Sun")
        date_str = day_entry["date"]

        # Date filtering: only count shifts within Block 3 boundaries
        d = parse_date(date_str)
        if not in_block3_range(d):
            dates_outside_range += 1
            continue

        for assignment in day_entry["assignments"]:
            # Clean provider name
            provider = clean_html_provider(assignment["provider"])
            if not provider:
                continue

            # Exclude moonlighting shifts
            if assignment.get("moonlighting", False):
                moonlighting_excluded += 1
                continue

            service_name = assignment["service"]
            svc_class = service_classes.get(service_name, "day")

            if svc_class == "exclude":
                continue

            # Determine shift type for this day
            if svc_class == "night":
                shift_type = "night"
            elif svc_class == "swing":
                shift_type = "swing"
            elif is_weekend:
                shift_type = "weekend"
            else:
                shift_type = "weekday"

            # Keep highest-priority shift type for this provider+date
            existing = provider_day_shifts[provider].get(date_str)
            priority = {"night": 3, "swing": 2, "weekday": 1, "weekend": 1}
            if existing is None or priority.get(shift_type, 0) > priority.get(existing, 0):
                provider_day_shifts[provider][date_str] = shift_type

    print(f"\nDate filtering: {dates_outside_range} days outside Block 3 range skipped")
    print(f"Moonlighting shifts excluded: {moonlighting_excluded}")

    # Aggregate from deduplicated day-level data
    provider_counts = defaultdict(lambda: {
        "weekday_shifts": 0,
        "weekend_shifts": 0,
        "night_shifts": 0,
        "swing_shifts": 0,
    })

    for provider, day_map in provider_day_shifts.items():
        for date_str, shift_type in day_map.items():
            if shift_type == "night":
                provider_counts[provider]["night_shifts"] += 1
            elif shift_type == "swing":
                provider_counts[provider]["swing_shifts"] += 1
            elif shift_type == "weekend":
                provider_counts[provider]["weekend_shifts"] += 1
            else:
                provider_counts[provider]["weekday_shifts"] += 1

    # Calculate weeks/weekends/nights using CANONICAL names
    results = {}
    for provider in sorted(provider_counts.keys()):
        counts = provider_counts[provider]
        canonical = to_canonical(provider)
        # If multiple HTML names map to the same canonical, sum them
        if canonical in results:
            results[canonical]["weekday_shifts"] += counts["weekday_shifts"]
            results[canonical]["weekend_shifts"] += counts["weekend_shifts"]
            results[canonical]["night_shifts"] += counts["night_shifts"]
            results[canonical]["swing_shifts"] += counts["swing_shifts"]
            results[canonical]["block3_weeks"] = round(
                results[canonical]["weekday_shifts"] / 5, 1)
            results[canonical]["block3_weekends"] = round(
                results[canonical]["weekend_shifts"] / 2, 1)
            results[canonical]["block3_nights"] = results[canonical]["night_shifts"]
        else:
            results[canonical] = {
                "weekday_shifts": counts["weekday_shifts"],
                "block3_weeks": round(counts["weekday_shifts"] / 5, 1),
                "weekend_shifts": counts["weekend_shifts"],
                "block3_weekends": round(counts["weekend_shifts"] / 2, 1),
                "night_shifts": counts["night_shifts"],
                "block3_nights": counts["night_shifts"],
                "swing_shifts": counts["swing_shifts"],
            }

    return results


# ---------------------------------------------------------------------------
# Output
# ---------------------------------------------------------------------------

def write_xlsx(provider_order, provider_data, prior_actuals, block3_actuals):
    """Write comparison spreadsheet matching Google Sheet provider order."""
    try:
        import openpyxl
    except ImportError:
        print("\nWARNING: openpyxl not installed — skipping XLSX output")
        print("  Install with: pip install openpyxl")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Block 3 Actuals"

    # Header row
    headers = [
        "provider_name",
        "annual_weeks",
        "annual_weekends",
        "annual_nights",
        "prior_weeks_worked",
        "prior_weekends_worked",
        "prior_nights_worked",
        "block3_weeks",
        "block3_weekends",
        "block3_nights",
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(1, col, header)

    row = 2
    matched_b3 = 0
    for sheet_name in provider_order:
        # Annual targets from providers sheet
        pdata = provider_data.get(sheet_name, {})
        ann_wk = pdata.get("annual_weeks", 0)
        ann_we = pdata.get("annual_weekends", 0)
        ann_nt = pdata.get("annual_nights", 0)

        # Prior actuals (Block 1&2)
        prior_key = match_provider(sheet_name, prior_actuals.keys())
        if prior_key:
            pa = prior_actuals[prior_key]
            prior_wk = pa.get("prior_weeks", 0)
            prior_we = pa.get("prior_weekends", 0)
            prior_nt = pa.get("prior_nights", 0)
        else:
            prior_wk = prior_we = prior_nt = 0

        # Block 3 actuals
        b3_key = match_provider(sheet_name, block3_actuals.keys())
        if b3_key:
            b3 = block3_actuals[b3_key]
            b3_wk = b3["block3_weeks"]
            b3_we = b3["block3_weekends"]
            b3_nt = b3["block3_nights"]
            matched_b3 += 1
        else:
            b3_wk = b3_we = b3_nt = 0

        ws.cell(row, 1, sheet_name)
        ws.cell(row, 2, ann_wk)
        ws.cell(row, 3, ann_we)
        ws.cell(row, 4, ann_nt)
        ws.cell(row, 5, prior_wk)
        ws.cell(row, 6, prior_we)
        ws.cell(row, 7, prior_nt)
        ws.cell(row, 8, b3_wk)
        ws.cell(row, 9, b3_we)
        ws.cell(row, 10, b3_nt)
        row += 1

    xlsx_path = os.path.join(OUTPUT_DIR, "block3_actuals.xlsx")
    wb.save(xlsx_path)
    print(f"\nWrote XLSX: {xlsx_path}")
    print(f"  {len(provider_order)} providers, {matched_b3} with Block 3 shifts")


def print_report(provider_order, provider_data, prior_actuals, block3_actuals):
    """Print terminal summary table."""
    print(f"\n{'='*110}")
    print(f"{'Provider':<40s} {'AnnWk':>6s} {'AnnWE':>6s} {'AnnNt':>6s} "
          f"{'PrWk':>6s} {'PrWE':>6s} {'PrNt':>6s} "
          f"{'B3Wk':>6s} {'B3WE':>6s} {'B3Nt':>6s}")
    print("-" * 110)

    active_count = 0
    for sheet_name in provider_order:
        pdata = provider_data.get(sheet_name, {})
        ann_wk = pdata.get("annual_weeks", 0)
        ann_we = pdata.get("annual_weekends", 0)
        ann_nt = pdata.get("annual_nights", 0)

        prior_key = match_provider(sheet_name, prior_actuals.keys())
        if prior_key:
            pa = prior_actuals[prior_key]
            prior_wk = pa.get("prior_weeks", 0)
            prior_we = pa.get("prior_weekends", 0)
            prior_nt = pa.get("prior_nights", 0)
        else:
            prior_wk = prior_we = prior_nt = 0

        b3_key = match_provider(sheet_name, block3_actuals.keys())
        if b3_key:
            b3 = block3_actuals[b3_key]
            b3_wk = b3["block3_weeks"]
            b3_we = b3["block3_weekends"]
            b3_nt = b3["block3_nights"]
        else:
            b3_wk = b3_we = b3_nt = 0

        # Skip providers with zero across the board
        total = ann_wk + ann_we + ann_nt + prior_wk + prior_we + prior_nt + b3_wk + b3_we + b3_nt
        if total == 0:
            continue

        active_count += 1
        print(f"{sheet_name:<40s} {ann_wk:>6.0f} {ann_we:>6.0f} {ann_nt:>6.0f} "
              f"{prior_wk:>6.1f} {prior_we:>6.1f} {prior_nt:>6.0f} "
              f"{b3_wk:>6.1f} {b3_we:>6.1f} {b3_nt:>6.0f}")

    print(f"\n{active_count} active providers shown")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print("=" * 70)
    print("BLOCK 3 ACTUALS — PARSED FROM AMION HTML SCHEDULES")
    print(f"Block 3: {BLOCK_3_START} to {BLOCK_3_END}")
    print("=" * 70)

    # Load providers sheet (annual targets + provider order)
    print("\n--- Loading providers sheet ---")
    provider_order, provider_data = load_providers_sheet()

    # Load prior actuals (Block 1&2)
    print("\n--- Loading Block 1&2 prior actuals ---")
    prior_path = os.path.join(OUTPUT_DIR, "prior_actuals.json")
    if os.path.isfile(prior_path):
        with open(prior_path) as f:
            prior_actuals = json.load(f)
        print(f"  Loaded {len(prior_actuals)} providers from prior_actuals.json")
    else:
        print(f"  WARNING: {prior_path} not found — Block 1&2 columns will be zeros")
        prior_actuals = {}

    # Compute Block 3 actuals from HTML schedules
    print("\n--- Parsing Block 3 schedules ---")
    block3_actuals = compute_block3_actuals()
    print(f"\n{len(block3_actuals)} providers found in Block 3 schedules")

    # Write outputs
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    write_xlsx(provider_order, provider_data, prior_actuals, block3_actuals)
    print_report(provider_order, provider_data, prior_actuals, block3_actuals)


if __name__ == "__main__":
    main()
