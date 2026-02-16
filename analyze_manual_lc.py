#!/usr/bin/env python3
"""
Analyze manual long call assignments against the full work schedules.
Cross-references HTML schedule data with Excel LC assignments for all 3 blocks.
Checks all strong/soft rules and identifies violations vs patterns.
"""

import os
import sys
import json
import glob
import re
import openpyxl
from datetime import datetime, timedelta
from collections import defaultdict

# Add project root to path
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, SCRIPT_DIR)
from parse_schedule import parse_schedule, merge_schedules

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
INPUT_DIR = os.path.join(SCRIPT_DIR, "input")
CONFIG_PATH = os.path.join(SCRIPT_DIR, "config.json")
EXCEL_PATH = os.path.join(INPUT_DIR, "Long call 2025-26.xlsx")

with open(CONFIG_PATH) as f:
    CONFIG = json.load(f)

TEACHING_SERVICES = set(CONFIG["teaching_services"])
DC_SERVICES = set(CONFIG["direct_care_services"])
SOURCE_SERVICES = TEACHING_SERVICES | DC_SERVICES

# Holiday dates (from config + known holidays for all blocks)
HOLIDAYS = set()
for h in CONFIG.get("holidays", []):
    HOLIDAYS.add(datetime.strptime(h, "%Y-%m-%d").date())
# Add known holidays that span all blocks
KNOWN_HOLIDAYS = [
    "2025-07-04",  # Independence Day
    "2025-09-01",  # Labor Day
    "2025-11-27",  # Thanksgiving
    "2025-12-25",  # Christmas
    "2026-01-01",  # New Year's
    "2026-05-25",  # Memorial Day
]
for h in KNOWN_HOLIDAYS:
    HOLIDAYS.add(datetime.strptime(h, "%Y-%m-%d").date())

BLOCK_CONFIGS = {
    "Block 1": {
        "sheet": "Block 1 2025-26",
        "months": ["6_1 to 6_30, 2025", "7_1 to 7_31, 2025", "8_1 to 8_31, 2025",
                    "9_1 to 9_30, 2025", "10_1 to 10_31, 2025"],
    },
    "Block 2": {
        "sheet": "Block 2 2025-26",
        "months": ["11_1 to 11_30, 2025", "12_1 to 12_31, 2025",
                    "1_1 to 1_31, 2026", "2_1 to 2_28, 2026"],
    },
    "Block 3": {
        "sheet": "Block 3 2025-26",
        "months": ["3_1 to 3_31, 2026", "4_1 to 4_30, 2026",
                    "5_1 to 5_31, 2026", "6_1 to 6_30, 2026"],
    },
}

# Manual name aliases for Excel names that fuzzy matching can't resolve.
# Maps Excel name -> target "Last, First" format used in HTML schedules.
NAME_ALIASES = {
    "Kiki Li": "Li, Ka Yi",
    # Block 1 typos
    "Jasjit Dhilon": "Dhillon, Jasjit",
    "Jon Edelsetin": "Edelstein, Jonathan",
    "Marjorie Cadstin": "Cadestin, Marjorie",
    "Sadia Naswhin": "Nawshin, Sadia",
    # Block 2 typos
    "Adi Sapassetty": "Sapasetty , Aditya",
    "Cindy Glickman": "Glickman, Cynthia",
    "Vincent Maioriano": "Maiorino, Vincent",
    # Block 3 typos
    "Angela Zhang": "Zheng, Angela",
    "Chris Fernanez": "Fernandez, Christopher",
    "Gabriela Contrino": "Contino, Gabriela",
    "Lliang Xiaohui": "Liang, Xiaohui",
    "Paul Mcmackini": "McMackin, Paul",
}


# ---------------------------------------------------------------------------
# Parse Excel LC data
# ---------------------------------------------------------------------------
def parse_excel_lc():
    """Parse all 3 block sheets + attending from the Excel file."""
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    blocks = {}

    for block_name, cfg in BLOCK_CONFIGS.items():
        ws = wb[cfg["sheet"]]
        assignments = []  # list of {date, teaching, dc1, dc2, no_call}
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            date_val = row[0]
            if not date_val or not hasattr(date_val, "weekday"):
                continue
            assignments.append({
                "date": date_val.date() if hasattr(date_val, "date") else date_val,
                "teaching": str(row[1]).strip() if row[1] else None,
                "dc1": str(row[2]).strip() if row[2] else None,
                "dc2": str(row[3]).strip() if row[3] else None,
                "no_call": str(row[4]).strip() if row[4] else None,
            })
        blocks[block_name] = assignments

    # Attending sheet
    ws_att = wb["attending"]
    attending = []
    for row in ws_att.iter_rows(min_row=2, max_row=ws_att.max_row, values_only=True):
        name = row[0]
        if not name or name == "Attending":
            continue
        att = {"name": str(name).strip()}
        for i, key in enumerate(["we_b1", "we_b2", "we_b3", "dbl_b1", "dbl_b2", "dbl_b3",
                                  "hol_2023", "hol_2024", "hol_2025"]):
            val = row[i + 1] if i + 1 < len(row) else None
            if val and hasattr(val, "strftime"):
                att[key] = val.date() if hasattr(val, "date") else val
            elif val:
                att[key] = str(val).strip()
            else:
                att[key] = None
        attending.append(att)

    return blocks, attending


# ---------------------------------------------------------------------------
# Parse HTML schedules
# ---------------------------------------------------------------------------
def parse_html_schedules():
    """Parse all HTML schedule files and build a date->provider->services map."""
    all_files = sorted(glob.glob(os.path.join(INPUT_DIR, "Hospital Medicine Schedule*.html")))
    all_month_data = []
    for f in all_files:
        month_data = parse_schedule(f)
        all_month_data.append(month_data)

    merged = merge_schedules(all_month_data)

    # Build: date_str -> [{service, provider, moonlighting, note}]
    schedule_by_date = defaultdict(list)
    for day in merged["schedule"]:
        date_str = day["date"]
        for a in day["assignments"]:
            if a["provider"] and a["provider"] not in ("", "-", "OPEN SHIFT"):
                schedule_by_date[date_str].append({
                    "service": a["service"],
                    "provider": a["provider"],
                    "moonlighting": a["moonlighting"],
                    "note": a.get("note", ""),
                })
    return schedule_by_date, merged


def normalize_date_key(dt):
    """Convert a date object to the M/D/YYYY format used by the parser."""
    return f"{dt.month}/{dt.day}/{dt.year}"


# ---------------------------------------------------------------------------
# Build provider stretches from schedule data
# ---------------------------------------------------------------------------
def build_provider_work_days(schedule_by_date, block_dates):
    """For each provider, find the dates they work on source services within the block."""
    provider_days = defaultdict(lambda: defaultdict(set))  # provider -> date -> set of services

    for dt in block_dates:
        key = normalize_date_key(dt)
        for entry in schedule_by_date.get(key, []):
            if entry["service"] in SOURCE_SERVICES and not entry.get("moonlighting", False):
                provider_days[entry["provider"]][dt].add(entry["service"])

    return provider_days


def build_stretches(provider_days):
    """Build consecutive work stretches for each provider."""
    stretches = {}  # provider -> list of [date, date, ...]
    for provider, day_map in provider_days.items():
        sorted_dates = sorted(day_map.keys())
        if not sorted_dates:
            continue
        current_stretch = [sorted_dates[0]]
        all_stretches = []
        for i in range(1, len(sorted_dates)):
            if (sorted_dates[i] - sorted_dates[i - 1]).days == 1:
                current_stretch.append(sorted_dates[i])
            else:
                all_stretches.append(current_stretch)
                current_stretch = [sorted_dates[i]]
        all_stretches.append(current_stretch)
        stretches[provider] = all_stretches
    return stretches


def get_provider_category(provider, date, schedule_by_date):
    """Determine if a provider is teaching or DC on a given date."""
    key = normalize_date_key(date)
    for entry in schedule_by_date.get(key, []):
        if entry["provider"] == provider:
            if entry["service"] in TEACHING_SERVICES:
                return "teaching"
            elif entry["service"] in DC_SERVICES:
                return "dc"
    return None


def is_weekend_or_holiday(dt):
    """Check if a date is a weekend or holiday."""
    if isinstance(dt, datetime):
        dt = dt.date()
    return dt.weekday() >= 5 or dt in HOLIDAYS


def is_next_iso_week(wk1, wk2):
    """Check if wk2 is the ISO week immediately following wk1.

    Handles year boundaries (e.g., (2025, 52) -> (2026, 1)).
    wk1 and wk2 are (year, week_number) tuples.
    """
    y1, w1 = wk1
    y2, w2 = wk2
    # Same year, next week
    if y1 == y2 and w2 == w1 + 1:
        return True
    # Year boundary: last week of y1 -> week 1 of y1+1
    if y2 == y1 + 1 and w2 == 1:
        # ISO years can have 52 or 53 weeks
        # Check if w1 is the last week of its year
        from datetime import date
        dec28 = date(y1, 12, 28)  # Dec 28 is always in the last ISO week
        last_week = dec28.isocalendar()[1]
        if w1 == last_week:
            return True
    return False


def slots_for_date(dt):
    """Return the number of LC slots available on a given date.

    Weekdays: 3 (teaching + dc1 + dc2)
    Weekends/holidays: 2 (teaching + dc2)
    """
    if is_weekend_or_holiday(dt):
        return 2
    return 3


# ---------------------------------------------------------------------------
# Detect swaps
# ---------------------------------------------------------------------------
def find_swaps(schedule_by_date, block_dates):
    """Find all swap notes on source services within block dates."""
    swaps = []
    swap_kw = ["swap", "switch", "cover", "payback", "paid back"]
    for dt in block_dates:
        key = normalize_date_key(dt)
        for entry in schedule_by_date.get(key, []):
            if entry["service"] in SOURCE_SERVICES and entry["note"]:
                note_lower = entry["note"].lower()
                if any(kw in note_lower for kw in swap_kw):
                    swaps.append({
                        "date": dt,
                        "provider": entry["provider"],
                        "service": entry["service"],
                        "note": entry["note"],
                    })
    return swaps


# ---------------------------------------------------------------------------
# Name matching (Excel uses first names, HTML uses Last, First)
# ---------------------------------------------------------------------------
def build_name_index(schedule_by_date, block_dates):
    """Build index of all provider names seen in schedules for fuzzy matching."""
    names = set()
    for dt in block_dates:
        key = normalize_date_key(dt)
        for entry in schedule_by_date.get(key, []):
            if entry["service"] in SOURCE_SERVICES:
                names.add(entry["provider"])
    return names


def match_excel_name(excel_name, html_names, _cache={}):
    """Match an Excel name (e.g. 'Emily Cunnings') to HTML format ('Cunnings, Emily')."""
    if excel_name in _cache:
        return _cache[excel_name]

    if not excel_name:
        _cache[excel_name] = None
        return None

    # Clean up
    name = excel_name.strip()

    # Check manual aliases first (for nicknames like "Kiki Li" -> "Li, Ka Yi")
    if name in NAME_ALIASES:
        _cache[excel_name] = NAME_ALIASES[name]
        return NAME_ALIASES[name]

    # Try direct match first
    if name in html_names:
        _cache[name] = name
        return name

    # Try "Last, First" format
    parts = name.split()
    if len(parts) >= 2:
        # Try last-first
        last = parts[-1]
        first = " ".join(parts[:-1])
        candidate = f"{last}, {first}"
        if candidate in html_names:
            _cache[excel_name] = candidate
            return candidate

        # Try first-last reversed
        first = parts[0]
        last = " ".join(parts[1:])
        candidate = f"{last}, {first}"
        if candidate in html_names:
            _cache[excel_name] = candidate
            return candidate

    # Fuzzy: find names containing key parts
    name_parts = name.lower().split()
    for html_name in html_names:
        html_lower = html_name.lower()
        if all(part in html_lower for part in name_parts):
            _cache[excel_name] = html_name
            return html_name

    # Last resort: match on last name only
    if len(parts) >= 2:
        last_name = parts[-1].lower()
        matches = [n for n in html_names if n.lower().startswith(last_name + ",") or
                   n.lower().startswith(last_name + " ")]
        if len(matches) == 1:
            _cache[excel_name] = matches[0]
            return matches[0]

    _cache[excel_name] = None
    return None


# ---------------------------------------------------------------------------
# Analysis functions
# ---------------------------------------------------------------------------
def analyze_block(block_name, lc_assignments, schedule_by_date, attending):
    """Run full analysis on one block."""
    print(f"\n{'='*70}")
    print(f"  {block_name}")
    print(f"{'='*70}")

    if not lc_assignments:
        print("  No LC assignments found.")
        return {}

    block_dates = [a["date"] for a in lc_assignments]
    date_range = f"{block_dates[0]} to {block_dates[-1]}"
    print(f"  Date range: {date_range} ({len(block_dates)} days)")

    # Build name index
    html_names = build_name_index(schedule_by_date, block_dates)

    # Build provider work data
    provider_days = build_provider_work_days(schedule_by_date, block_dates)
    stretches = build_stretches(provider_days)

    # Build LC assignment map: provider -> [(date, slot_type)]
    provider_lcs = defaultdict(list)
    provider_weekend_lcs = defaultdict(int)
    lc_by_date = {}  # date -> {teaching, dc1, dc2}
    unmatched_names = set()

    for a in lc_assignments:
        dt = a["date"]
        day_lcs = {}
        for slot in ["teaching", "dc1", "dc2"]:
            excel_name = a[slot]
            if not excel_name or excel_name.lower() == "none":
                continue
            html_name = match_excel_name(excel_name, html_names)
            if html_name:
                provider_lcs[html_name].append((dt, slot))
                if is_weekend_or_holiday(dt):
                    provider_weekend_lcs[html_name] += 1
                day_lcs[slot] = html_name
            else:
                unmatched_names.add(excel_name)
                day_lcs[slot] = excel_name  # keep raw name
                provider_lcs[excel_name].append((dt, slot))
                if is_weekend_or_holiday(dt):
                    provider_weekend_lcs[excel_name] += 1
        lc_by_date[dt] = day_lcs

    if unmatched_names:
        print(f"\n  ⚠ Unmatched Excel names ({len(unmatched_names)}):")
        for n in sorted(unmatched_names):
            print(f"    - {n}")

    # Find swaps
    swaps = find_swaps(schedule_by_date, block_dates)
    swap_dates = {(s["date"], s["provider"]) for s in swaps}

    results = {
        "block": block_name,
        "date_range": date_range,
        "total_days": len(block_dates),
        "total_providers": len(provider_lcs),
        "violations": [],
        "soft_stats": {},
    }

    # -------------------------------------------------------------------
    # STRONG RULE CHECKS
    # -------------------------------------------------------------------
    print(f"\n  --- Strong Rule Checks ---")

    # Rule 1: DC provider on teaching slot
    rule1_violations = []
    for a in lc_assignments:
        dt = a["date"]
        if a["teaching"]:
            excel_name = a["teaching"]
            html_name = match_excel_name(excel_name, html_names)
            name_to_check = html_name or excel_name
            cat = get_provider_category(name_to_check, dt, schedule_by_date)
            if cat == "dc":
                is_we = is_weekend_or_holiday(dt)
                rule1_violations.append({
                    "date": dt, "provider": name_to_check,
                    "weekend": is_we,
                    "swap": (dt, name_to_check) in swap_dates,
                })
    print(f"  Rule 1 (DC on teaching): {len(rule1_violations)} violations")
    for v in rule1_violations:
        tag = " [SWAP]" if v["swap"] else ""
        we = " (weekend)" if v["weekend"] else " (weekday)"
        print(f"    {v['date']} {v['provider']}{we}{tag}")
    results["violations"].append(("Rule 1", rule1_violations))

    # -------------------------------------------------------------------
    # ANALYSIS 1: Rule 1 vs Rule 2 Cross-Check (reciprocal overflow)
    # -------------------------------------------------------------------
    print(f"\n  --- Analysis 1: Rule 1 vs Rule 2 Cross-Check ---")

    # Group Rule 1 violations by ISO week
    rule1_by_week = defaultdict(list)
    for v in rule1_violations:
        iso = v["date"].isocalendar()
        rule1_by_week[(iso[0], iso[1])].append(v)

    # Find all Rule 2 overflow: teaching providers assigned to DC LC slots
    rule2_overflow = []
    for a in lc_assignments:
        dt = a["date"]
        iso = dt.isocalendar()
        wk = (iso[0], iso[1])
        for slot in ["dc1", "dc2"]:
            excel_name = a[slot]
            if not excel_name or excel_name.lower() == "none":
                continue
            html_name = match_excel_name(excel_name, html_names)
            name_to_check = html_name or excel_name
            cat = get_provider_category(name_to_check, dt, schedule_by_date)
            if cat == "teaching":
                rule2_overflow.append({
                    "date": dt,
                    "provider": name_to_check,
                    "slot": slot,
                    "week": wk,
                })

    # Group Rule 2 overflow by ISO week
    rule2_by_week = defaultdict(list)
    for o in rule2_overflow:
        rule2_by_week[o["week"]].append(o)

    # Cross-reference: weeks with both Rule 1 AND Rule 2
    reciprocal_weeks = []
    for wk in sorted(rule1_by_week.keys()):
        r1_list = rule1_by_week[wk]
        r2_list = rule2_by_week.get(wk, [])
        reciprocal_weeks.append({
            "week": wk,
            "rule1_count": len(r1_list),
            "rule2_count": len(r2_list),
            "rule1_details": r1_list,
            "rule2_details": r2_list,
            "is_reciprocal": len(r2_list) > 0,
        })

    reciprocal_count = sum(1 for rw in reciprocal_weeks if rw["is_reciprocal"])
    print(f"  Rule 1 violation weeks: {len(reciprocal_weeks)}")
    print(f"  Weeks with reciprocal Rule 2 overflow: {reciprocal_count}")
    print(f"  Weeks without reciprocal: {len(reciprocal_weeks) - reciprocal_count}")
    print(f"  Total Rule 2 overflow instances (block-wide): {len(rule2_overflow)}")

    for rw in reciprocal_weeks:
        wk_label = f"Week {rw['week'][1]} ({rw['week'][0]})"
        if rw["is_reciprocal"]:
            print(f"\n    {wk_label}: RECIPROCAL")
            print(f"      Rule 1 (DC->teaching LC):")
            for v in rw["rule1_details"]:
                tag = " [SWAP]" if v["swap"] else ""
                we = " (weekend)" if v["weekend"] else ""
                print(f"        {v['date']} {v['provider']}{we}{tag}")
            print(f"      Rule 2 (teaching->DC LC):")
            for o in rw["rule2_details"]:
                print(f"        {o['date']} {o['provider']} ({o['slot']})")
        else:
            print(f"\n    {wk_label}: NO RECIPROCAL")
            for v in rw["rule1_details"]:
                tag = " [SWAP]" if v["swap"] else ""
                we = " (weekend)" if v["weekend"] else ""
                print(f"        {v['date']} {v['provider']}{we}{tag}")

    # Standalone Rule 2 weeks (overflow with no Rule 1 violations)
    standalone_r2_weeks = sorted([wk for wk in rule2_by_week if wk not in rule1_by_week])
    print(f"\n  Rule 2 overflow weeks with NO Rule 1 violations: {len(standalone_r2_weeks)}")

    # Weekend vs weekday breakdown for Rule 1
    r1_weekday = [v for v in rule1_violations if not v["weekend"]]
    r1_weekend = [v for v in rule1_violations if v["weekend"]]
    print(f"\n  Rule 1 breakdown: {len(r1_weekday)} weekday, {len(r1_weekend)} weekend")

    # Same-day reciprocal check: on the specific day of a Rule 1 violation,
    # was a teaching provider also assigned to a DC slot?
    rule2_by_date = defaultdict(list)
    for o in rule2_overflow:
        rule2_by_date[o["date"]].append(o)

    r1_sameday_recip = 0
    r1_sameday_no_recip = 0
    r1_weekend_sameday_recip = 0
    r1_weekend_sameday_no_recip = 0
    print(f"\n  Same-day reciprocal check (Rule 1 + Rule 2 on exact same date):")
    for v in rule1_violations:
        r2_same = rule2_by_date.get(v["date"], [])
        we_tag = " (weekend)" if v["weekend"] else " (weekday)"
        if r2_same:
            r1_sameday_recip += 1
            if v["weekend"]:
                r1_weekend_sameday_recip += 1
            r2_names = ", ".join(f"{o['provider']}({o['slot']})" for o in r2_same)
            print(f"    {v['date']} {v['provider']}{we_tag}: YES -> {r2_names}")
        else:
            r1_sameday_no_recip += 1
            if v["weekend"]:
                r1_weekend_sameday_no_recip += 1
            print(f"    {v['date']} {v['provider']}{we_tag}: no same-day Rule 2")

    print(f"\n  Same-day reciprocal summary:")
    print(f"    All Rule 1: {r1_sameday_recip} same-day reciprocal, "
          f"{r1_sameday_no_recip} not same-day")
    if r1_weekend:
        print(f"    Weekend Rule 1 only: {r1_weekend_sameday_recip} same-day reciprocal, "
              f"{r1_weekend_sameday_no_recip} not same-day "
              f"(out of {len(r1_weekend)} weekend violations)")

    results["analysis1_crosscheck"] = {
        "reciprocal_weeks": reciprocal_count,
        "non_reciprocal_weeks": len(reciprocal_weeks) - reciprocal_count,
        "total_rule2_overflow": len(rule2_overflow),
        "standalone_rule2_weeks": len(standalone_r2_weeks),
        "rule1_weekday": len(r1_weekday),
        "rule1_weekend": len(r1_weekend),
        "sameday_reciprocal": r1_sameday_recip,
        "sameday_not_reciprocal": r1_sameday_no_recip,
        "weekend_sameday_reciprocal": r1_weekend_sameday_recip,
        "weekend_sameday_not_reciprocal": r1_weekend_sameday_no_recip,
    }

    # -------------------------------------------------------------------
    # ANALYSIS 2: Provider Crossover Patterns
    # -------------------------------------------------------------------
    print(f"\n  --- Analysis 2: Provider Crossover Patterns ---")

    rule1_providers = set(v["provider"] for v in rule1_violations)

    if not rule1_providers:
        print("  No Rule 1 violations — skipping crossover analysis.")
        results["analysis2_crossover"] = {"rule1_providers": 0}
    else:
        crossover_profiles = []
        for provider in sorted(rule1_providers):
            day_map = provider_days.get(provider, {})
            teaching_days = 0
            dc_days = 0
            both_days = 0
            total_days = len(day_map)

            for dt, services in day_map.items():
                has_teaching = bool(services & TEACHING_SERVICES)
                has_dc = bool(services & DC_SERVICES)
                if has_teaching and has_dc:
                    both_days += 1
                    teaching_days += 1
                    dc_days += 1
                elif has_teaching:
                    teaching_days += 1
                elif has_dc:
                    dc_days += 1

            if total_days == 0:
                category = "unknown (no schedule data)"
                teaching_pct = 0.0
                dc_pct = 0.0
            else:
                teaching_pct = teaching_days / total_days * 100
                dc_pct = dc_days / total_days * 100
                if teaching_pct >= 80:
                    category = "primarily teaching"
                elif dc_pct >= 80:
                    category = "primarily DC"
                else:
                    category = "crossover"

            r1_count = sum(1 for v in rule1_violations if v["provider"] == provider)

            crossover_profiles.append({
                "provider": provider,
                "total_days": total_days,
                "teaching_days": teaching_days,
                "dc_days": dc_days,
                "both_days": both_days,
                "teaching_pct": round(teaching_pct, 1),
                "dc_pct": round(dc_pct, 1),
                "category": category,
                "rule1_count": r1_count,
            })

        print(f"  Rule 1 violator profiles ({len(crossover_profiles)} providers):")
        for p in crossover_profiles:
            print(f"\n    {p['provider']}:")
            print(f"      Total work days: {p['total_days']}")
            print(f"      Teaching days: {p['teaching_days']} ({p['teaching_pct']}%)")
            print(f"      DC days: {p['dc_days']} ({p['dc_pct']}%)")
            if p["both_days"] > 0:
                print(f"      Both on same day: {p['both_days']}")
            print(f"      Category: {p['category']}")
            print(f"      Rule 1 violations: {p['rule1_count']}")

        cat_counts = defaultdict(int)
        for p in crossover_profiles:
            cat_counts[p["category"]] += 1

        print(f"\n  Category distribution of Rule 1 violators:")
        for cat in ["primarily teaching", "crossover", "primarily DC", "unknown (no schedule data)"]:
            if cat in cat_counts:
                print(f"    {cat}: {cat_counts[cat]}")

        # Compare against non-violating providers
        all_lc_providers = set(provider_lcs.keys())
        non_rule1_providers = all_lc_providers - rule1_providers
        non_rule1_crossover_count = 0
        non_rule1_total = 0
        for provider in non_rule1_providers:
            day_map = provider_days.get(provider, {})
            total = len(day_map)
            if total == 0:
                continue
            non_rule1_total += 1
            t_days = sum(1 for dt, svcs in day_map.items() if svcs & TEACHING_SERVICES)
            d_days = sum(1 for dt, svcs in day_map.items() if svcs & DC_SERVICES)
            t_pct = t_days / total * 100
            d_pct = d_days / total * 100
            if t_pct < 80 and d_pct < 80:
                non_rule1_crossover_count += 1

        rule1_crossover_count = cat_counts.get("crossover", 0)
        rule1_with_data = len(crossover_profiles) - cat_counts.get("unknown (no schedule data)", 0)

        print(f"\n  Crossover rate comparison:")
        if rule1_with_data > 0:
            print(f"    Rule 1 violators: {rule1_crossover_count}/{rule1_with_data} "
                  f"({rule1_crossover_count/rule1_with_data*100:.0f}%) are crossover")
        if non_rule1_total > 0:
            print(f"    Non-violators:    {non_rule1_crossover_count}/{non_rule1_total} "
                  f"({non_rule1_crossover_count/non_rule1_total*100:.0f}%) are crossover")

        results["analysis2_crossover"] = {
            "rule1_providers": len(rule1_providers),
            "profiles": crossover_profiles,
            "category_distribution": dict(cat_counts),
            "non_violator_crossover_rate": (
                non_rule1_crossover_count / non_rule1_total * 100
                if non_rule1_total > 0 else 0
            ),
        }

    # Rule 3: Two weekday LCs in same ISO week
    rule3_violations = []
    for provider, lcs in provider_lcs.items():
        week_weekday_lcs = defaultdict(list)
        for dt, slot in lcs:
            if not is_weekend_or_holiday(dt):
                iso = dt.isocalendar()
                week_weekday_lcs[(iso[0], iso[1])].append((dt, slot))
        for wk, wk_lcs in week_weekday_lcs.items():
            if len(wk_lcs) >= 2:
                rule3_violations.append({
                    "provider": provider, "week": wk,
                    "lcs": [(dt.isoformat(), slot) for dt, slot in wk_lcs],
                })
    print(f"  Rule 3 (two weekday LCs in ISO week): {len(rule3_violations)} violations")
    for v in rule3_violations:
        lc_str = ", ".join(f"{d} ({s})" for d, s in v["lcs"])
        print(f"    {v['provider']}: week {v['week'][1]} -> {lc_str}")
    results["violations"].append(("Rule 3", rule3_violations))

    # Rule 7: Min 2-day gap between LCs in same stretch
    rule7_violations = []
    for provider, strs in stretches.items():
        for stretch in strs:
            stretch_set = set(stretch)
            lcs_in_stretch = [(dt, slot) for dt, slot in provider_lcs.get(provider, [])
                              if dt in stretch_set]
            lcs_in_stretch.sort()
            for i in range(1, len(lcs_in_stretch)):
                gap = (lcs_in_stretch[i][0] - lcs_in_stretch[i - 1][0]).days
                if gap < 2:
                    rule7_violations.append({
                        "provider": provider,
                        "lc1": (lcs_in_stretch[i - 1][0].isoformat(), lcs_in_stretch[i - 1][1]),
                        "lc2": (lcs_in_stretch[i][0].isoformat(), lcs_in_stretch[i][1]),
                        "gap": gap,
                    })
    print(f"  Rule 7 (min 2-day gap): {len(rule7_violations)} violations")
    for v in rule7_violations:
        print(f"    {v['provider']}: {v['lc1'][0]} ({v['lc1'][1]}) -> {v['lc2'][0]} ({v['lc2'][1]}), gap={v['gap']}")
    results["violations"].append(("Rule 7", rule7_violations))

    # Rule 8: No double if stretch has a holiday LC
    rule8_violations = []
    for provider, strs in stretches.items():
        for stretch in strs:
            stretch_set = set(stretch)
            lcs_in_stretch = [(dt, slot) for dt, slot in provider_lcs.get(provider, [])
                              if dt in stretch_set]
            if len(lcs_in_stretch) >= 2:
                has_holiday = any(dt in HOLIDAYS for dt, _ in lcs_in_stretch)
                if has_holiday:
                    rule8_violations.append({
                        "provider": provider,
                        "lcs": [(dt.isoformat(), slot) for dt, slot in lcs_in_stretch],
                        "stretch": f"{stretch[0].isoformat()} to {stretch[-1].isoformat()}",
                    })
    print(f"  Rule 8 (no double with holiday LC): {len(rule8_violations)} violations")
    for v in rule8_violations:
        lc_str = ", ".join(f"{d} ({s})" for d, s in v["lcs"])
        print(f"    {v['provider']}: {lc_str} in stretch {v['stretch']}")
    results["violations"].append(("Rule 8", rule8_violations))

    # Rule 9: No dc1 on weekends/holidays
    rule9_violations = []
    for a in lc_assignments:
        dt = a["date"]
        if is_weekend_or_holiday(dt) and a["dc1"]:
            rule9_violations.append({"date": dt, "provider": a["dc1"]})
    print(f"  Rule 9 (no dc1 on weekend/holiday): {len(rule9_violations)} violations")
    for v in rule9_violations:
        print(f"    {v['date']} {v['provider']}")
    results["violations"].append(("Rule 9", rule9_violations))

    # Rule 10: Max 2 weekend LCs per provider
    rule10_violations = []
    for provider, count in provider_weekend_lcs.items():
        if count > 2:
            rule10_violations.append({"provider": provider, "count": count})
    print(f"  Rule 10 (max 2 weekend LCs): {len(rule10_violations)} violations")
    for v in rule10_violations:
        print(f"    {v['provider']}: {v['count']} weekend LCs")
    results["violations"].append(("Rule 10", rule10_violations))

    # Rule 11: All slots filled before doubles (no miss + double in same week)
    # Corrected logic: a violation only occurs when someone gets a double in a week
    # where there are unfilled LC slots (total assigned < total available).
    rule11_violations = []

    # Group LCs by ISO week
    week_lcs = defaultdict(lambda: defaultdict(list))  # week -> provider -> [(date, slot)]
    for provider, lcs in provider_lcs.items():
        for dt, slot in lcs:
            iso = dt.isocalendar()
            week_lcs[(iso[0], iso[1])][provider].append((dt, slot))

    # Provider work weeks (needed for Rule 12 too)
    # Only count weeks where the provider works at least 1 weekday (Mon-Fri).
    # Weekend-only weeks (e.g. Sat/Sun overflow shifts) don't count as work weeks.
    provider_work_weeks = defaultdict(set)  # provider -> set of iso weeks
    for provider, day_map in provider_days.items():
        for dt in day_map:
            if dt.weekday() < 5:  # Mon=0 .. Fri=4
                iso = dt.isocalendar()
                provider_work_weeks[provider].add((iso[0], iso[1]))

    # Build block dates by ISO week for slot counting
    block_dates_set = set(block_dates)
    block_dates_by_week = defaultdict(list)  # iso_week -> [dates in block]
    for dt in block_dates:
        iso = dt.isocalendar()
        block_dates_by_week[(iso[0], iso[1])].append(dt)

    # Provider work days per week (for detecting "passed over" providers)
    provider_days_per_week = defaultdict(lambda: defaultdict(int))  # provider -> week -> num_days
    for provider, day_map in provider_days.items():
        for dt in day_map:
            iso = dt.isocalendar()
            provider_days_per_week[provider][(iso[0], iso[1])] += 1

    for wk, prov_lcs in week_lcs.items():
        # Who has doubles this week?
        doubles_this_week = {p for p, lcs in prov_lcs.items() if len(lcs) >= 2}
        if not doubles_this_week:
            continue

        # Count total available slots this week
        week_dates = block_dates_by_week.get(wk, [])
        total_available = sum(slots_for_date(dt) for dt in week_dates)

        # Count total LCs assigned this week
        total_assigned = sum(len(lcs) for lcs in prov_lcs.values())

        # Violation: double exists AND unfilled slots
        if total_assigned < total_available:
            unfilled = total_available - total_assigned
            # Also find providers who worked ≥4 days this week but got no LC
            passed_over = []
            for provider, weeks in provider_work_weeks.items():
                if wk in weeks and provider not in prov_lcs:
                    days_worked = provider_days_per_week[provider].get(wk, 0)
                    if days_worked >= 4:
                        passed_over.append(f"{provider}({days_worked}d)")
            rule11_violations.append({
                "week": wk,
                "doubles": list(doubles_this_week),
                "available": total_available,
                "assigned": total_assigned,
                "unfilled": unfilled,
                "passed_over": passed_over,
            })

    print(f"  Rule 11 (fill all before doubles): {len(rule11_violations)} violations")
    for v in rule11_violations:
        print(f"    Week {v['week'][1]}: doubles={v['doubles']}, "
              f"slots={v['available']}, assigned={v['assigned']}, "
              f"unfilled={v['unfilled']}")
        if v["passed_over"]:
            print(f"      Passed-over providers (≥4 workdays, 0 LCs): {v['passed_over']}")
    results["violations"].append(("Rule 11", rule11_violations))

    # Rule 12: No 2+ consecutive weeks without LC (sliding window approach)
    # Unit of work = weekday work week (Mon-Fri). Weekend-only weeks don't count.
    # Each weekday work week gets a "window" that includes:
    #   - The Mon-Fri days of that ISO week
    #   - The preceding Sat-Sun if the provider works them AND the prior ISO week
    #     is NOT itself a weekday work week (WE+Week pattern)
    #   - The following Sat-Sun of the same ISO week if the provider works them
    #     (Week+WE pattern)
    # An LC anywhere in the window covers that weekday work week.
    # For Week+WE+Week: the shared weekend appears in both windows. Process
    # weeks in order; once a weekend LC is consumed by one week, it still appears
    # in the next window (the LC date doesn't get removed — each window just checks
    # intersection). This means a single weekend LC can credit both adjacent weeks,
    # which is correct: the provider DID get an LC during that stretch.

    def prev_iso_week(wk):
        """Return the ISO week before wk."""
        y, w = wk
        if w > 1:
            return (y, w - 1)
        # Week 1 -> last week of previous year
        from datetime import date as _date
        dec28 = _date(y - 1, 12, 28)
        last_week = dec28.isocalendar()[1]
        return (y - 1, last_week)

    rule12_violations = []
    for provider in provider_work_weeks:
        sorted_weeks = sorted(provider_work_weeks[provider])
        if len(sorted_weeks) < 2:
            continue

        work_dates = set(provider_days[provider].keys())
        weekday_work_weeks_set = provider_work_weeks[provider]

        # Index work dates by ISO week for fast lookup
        dates_by_week = defaultdict(list)
        for dt in work_dates:
            iso = dt.isocalendar()
            dates_by_week[(iso[0], iso[1])].append(dt)

        # Build set of LC dates for this provider
        lc_dates = set()
        for dt, slot in provider_lcs.get(provider, []):
            lc_dates.add(dt)

        consecutive_misses = 0
        max_consecutive = 0
        miss_runs = []
        current_run = []

        for i, wk in enumerate(sorted_weeks):
            # Build window: all work dates in this ISO week (weekdays + weekends)
            window_dates = set(dates_by_week.get(wk, []))

            # Add preceding Sat-Sun (from prior ISO week) if the provider
            # works them AND the prior ISO week is NOT a weekday work week
            # (i.e., it's a WE+Week pattern, not Week+WE+Week).
            pwk = prev_iso_week(wk)
            if pwk not in weekday_work_weeks_set:
                for dt in dates_by_week.get(pwk, []):
                    if dt.weekday() >= 5:
                        window_dates.add(dt)

            # Check if any LC falls in window
            has_lc = bool(window_dates & lc_dates)

            if not has_lc:
                # Miss — check if consecutive with previous
                if consecutive_misses == 0:
                    consecutive_misses = 1
                    current_run = [wk]
                elif is_next_iso_week(sorted_weeks[i - 1], wk):
                    consecutive_misses += 1
                    current_run.append(wk)
                else:
                    if consecutive_misses >= 2:
                        miss_runs.append(list(current_run))
                    consecutive_misses = 1
                    current_run = [wk]
            else:
                if consecutive_misses >= 2:
                    miss_runs.append(list(current_run))
                consecutive_misses = 0
                current_run = []
            max_consecutive = max(max_consecutive, consecutive_misses)

        # Handle trailing run
        if consecutive_misses >= 2:
            miss_runs.append(list(current_run))

        if max_consecutive >= 2:
            rule12_violations.append({
                "provider": provider,
                "max_consecutive": max_consecutive,
                "miss_runs": miss_runs,
            })
    print(f"  Rule 12 (no 2+ consecutive missed weeks): {len(rule12_violations)} violations")
    for v in rule12_violations:
        run_parts = []
        for run in v["miss_runs"]:
            week_dates = []
            for y, w in run:
                dates = block_dates_by_week.get((y, w), [])
                weekdays = sorted(d for d in dates if d.weekday() < 5)
                if weekdays:
                    week_dates.append(f"{weekdays[0].strftime('%b %d')}-{weekdays[-1].strftime('%b %d')}")
                else:
                    week_dates.append(f"wk{w}")
            run_parts.append(" | ".join(week_dates))
        runs_str = "; ".join(run_parts)
        print(f"    {v['provider']}: max {v['max_consecutive']} consecutive ({runs_str})")
    results["violations"].append(("Rule 12", rule12_violations))

    # -------------------------------------------------------------------
    # SOFT RULE ANALYSIS
    # -------------------------------------------------------------------
    print(f"\n  --- Soft Rule Analysis ---")

    # Rule A: Gap distribution
    gaps = []
    for provider, strs in stretches.items():
        for stretch in strs:
            stretch_set = set(stretch)
            lcs_in_stretch = [(dt, slot) for dt, slot in provider_lcs.get(provider, [])
                              if dt in stretch_set]
            lcs_in_stretch.sort()
            for i in range(1, len(lcs_in_stretch)):
                gap = (lcs_in_stretch[i][0] - lcs_in_stretch[i - 1][0]).days
                gaps.append(gap)
    if gaps:
        gap_dist = defaultdict(int)
        for g in gaps:
            gap_dist[g] += 1
        print(f"  Rule A (gap distribution): {len(gaps)} LC pairs in stretches")
        for g in sorted(gap_dist.keys()):
            print(f"    {g}-day gap: {gap_dist[g]}")
    results["soft_stats"]["gap_distribution"] = dict(defaultdict(int, {g: gaps.count(g) for g in set(gaps)})) if gaps else {}

    # Rule B: Weekend LC distribution
    we_dist = defaultdict(int)
    for provider, count in provider_weekend_lcs.items():
        we_dist[count] += 1
    zero_we = len(provider_lcs) - len(provider_weekend_lcs)
    we_dist[0] = zero_we
    print(f"  Rule B (weekend LC distribution):")
    for c in sorted(we_dist.keys()):
        print(f"    {c} weekend LCs: {we_dist[c]} providers")
    results["soft_stats"]["weekend_distribution"] = dict(we_dist)

    # Rule C: LCs proportional to weeks worked
    print(f"  Rule C (LC count vs weeks worked):")
    proportionality = []
    for provider in provider_lcs:
        lc_count = len(provider_lcs[provider])
        weeks_worked = len(provider_work_weeks.get(provider, set()))
        diff = lc_count - weeks_worked
        if abs(diff) > 1:
            proportionality.append({"provider": provider, "lcs": lc_count,
                                    "weeks": weeks_worked, "diff": diff})
    print(f"    Providers with |LC - weeks| > 1: {len(proportionality)}")
    for p in sorted(proportionality, key=lambda x: abs(x["diff"]), reverse=True)[:10]:
        print(f"    {p['provider']}: {p['lcs']} LCs, {p['weeks']} weeks (diff={p['diff']:+d})")
    results["soft_stats"]["proportionality_outliers"] = len(proportionality)

    # Rule D: Day-of-week variety
    print(f"  Rule D (day-of-week variety):")
    dow_violations = []
    for provider, lcs in provider_lcs.items():
        dow_counts = defaultdict(int)
        for dt, slot in lcs:
            dow_counts[dt.strftime("%A")] += 1
        for day, count in dow_counts.items():
            if count >= 3:
                dow_violations.append({"provider": provider, "day": day, "count": count})
    print(f"    Providers with 3+ LCs on same weekday: {len(dow_violations)}")
    for v in sorted(dow_violations, key=lambda x: x["count"], reverse=True)[:10]:
        print(f"    {v['provider']}: {v['count']}x {v['day']}")
    results["soft_stats"]["dow_violations"] = len(dow_violations)

    # Rule E: DC1/DC2 balance
    print(f"  Rule E (DC1/DC2 balance):")
    dc_balance_issues = []
    for provider, lcs in provider_lcs.items():
        dc1_count = sum(1 for _, s in lcs if s == "dc1")
        dc2_count = sum(1 for _, s in lcs if s == "dc2")
        if dc1_count > 0 or dc2_count > 0:
            if abs(dc1_count - dc2_count) > 1:
                dc_balance_issues.append({"provider": provider, "dc1": dc1_count, "dc2": dc2_count})
    print(f"    Providers with |dc1 - dc2| > 1: {len(dc_balance_issues)}")
    for p in sorted(dc_balance_issues, key=lambda x: abs(x["dc1"] - x["dc2"]), reverse=True)[:10]:
        print(f"    {p['provider']}: dc1={p['dc1']}, dc2={p['dc2']}")
    results["soft_stats"]["dc_balance_issues"] = len(dc_balance_issues)

    # Rule F: Missed weeks compensated by doubles
    print(f"  Rule F (missed weeks vs doubles):")
    providers_with_miss = set()
    providers_with_double = set()
    for provider, weeks in provider_work_weeks.items():
        got_lc_weeks = {dt.isocalendar()[:2] for dt, _ in provider_lcs.get(provider, [])}
        for wk in weeks:
            if wk not in got_lc_weeks:
                providers_with_miss.add(provider)
        # Check for doubles
        week_lc_count = defaultdict(int)
        for dt, _ in provider_lcs.get(provider, []):
            iso = dt.isocalendar()
            week_lc_count[(iso[0], iso[1])] += 1
        for wk, count in week_lc_count.items():
            if count >= 2:
                providers_with_double.add(provider)

    both = providers_with_miss & providers_with_double
    miss_only = providers_with_miss - providers_with_double
    double_only = providers_with_double - providers_with_miss
    print(f"    Providers with missed week: {len(providers_with_miss)}")
    print(f"    Providers with double: {len(providers_with_double)}")
    print(f"    Both (miss + double): {len(both)}")
    print(f"    Miss only (no compensating double): {len(miss_only)}")
    if miss_only:
        for p in sorted(miss_only):
            print(f"      - {p}")
    print(f"    Double only (no miss to compensate): {len(double_only)}")
    if double_only:
        for p in sorted(double_only):
            print(f"      - {p}")
    results["soft_stats"]["miss_double_correlation"] = {
        "miss_count": len(providers_with_miss),
        "double_count": len(providers_with_double),
        "both": len(both),
        "miss_only": len(miss_only),
        "double_only": len(double_only),
    }

    # -------------------------------------------------------------------
    # SWAP ANALYSIS
    # -------------------------------------------------------------------
    print(f"\n  --- Swap Analysis ---")
    swaps = find_swaps(schedule_by_date, block_dates)
    print(f"  Swaps on source services: {len(swaps)}")
    for s in swaps:
        print(f"    {s['date']} {s['provider']} ({s['service']}): {s['note']}")
    results["swaps"] = len(swaps)

    # -------------------------------------------------------------------
    # SUMMARY STATS
    # -------------------------------------------------------------------
    print(f"\n  --- Summary ---")
    total_lcs = sum(len(lcs) for lcs in provider_lcs.values())
    total_teaching = sum(1 for lcs in provider_lcs.values() for _, s in lcs if s == "teaching")
    total_dc1 = sum(1 for lcs in provider_lcs.values() for _, s in lcs if s == "dc1")
    total_dc2 = sum(1 for lcs in provider_lcs.values() for _, s in lcs if s == "dc2")
    total_doubles = sum(1 for p, lcs in provider_lcs.items()
                        if any(len([l for l in lcs if l[0].isocalendar()[:2] == wk]) >= 2
                               for wk in {dt.isocalendar()[:2] for dt, _ in lcs}))
    total_missed = len(providers_with_miss)

    print(f"  Total LCs: {total_lcs} (teaching={total_teaching}, dc1={total_dc1}, dc2={total_dc2})")
    print(f"  Providers with LCs: {len(provider_lcs)}")
    print(f"  Providers with doubles: {len(providers_with_double)}")
    print(f"  Providers with missed weeks: {total_missed}")
    print(f"  Weekend LCs total: {sum(provider_weekend_lcs.values())}")

    results["summary"] = {
        "total_lcs": total_lcs,
        "teaching": total_teaching,
        "dc1": total_dc1,
        "dc2": total_dc2,
        "providers": len(provider_lcs),
        "doubles": len(providers_with_double),
        "missed": total_missed,
        "weekend_lcs": sum(provider_weekend_lcs.values()),
    }

    return results


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    print("Parsing Excel LC data...")
    blocks, attending = parse_excel_lc()

    print("Parsing HTML schedules...")
    schedule_by_date, merged = parse_html_schedules()
    print(f"  Parsed {len(merged['schedule'])} days, {len(merged.get('by_provider', {}))} providers")

    all_results = {}
    for block_name in ["Block 1", "Block 2", "Block 3"]:
        lc_data = blocks[block_name]
        results = analyze_block(block_name, lc_data, schedule_by_date, attending)
        all_results[block_name] = results

    # Cross-block summary
    print(f"\n{'='*70}")
    print(f"  CROSS-BLOCK SUMMARY")
    print(f"{'='*70}")
    for block_name, results in all_results.items():
        if not results:
            continue
        s = results.get("summary", {})
        violations = results.get("violations", [])
        total_v = sum(len(v[1]) for v in violations)
        print(f"\n  {block_name}: {s.get('total_lcs', 0)} LCs, "
              f"{s.get('doubles', 0)} doubles, {s.get('missed', 0)} missed, "
              f"{total_v} strong violations")


if __name__ == "__main__":
    main()
