"""
Excel I/O for V3 Pre-Scheduler.

Reads from and writes to the hospitalist_scheduler.xlsx workbook.
Reader functions return the same data shapes as block/engines/shared/loader.py
so the engine can swap data sources without code changes.
"""

import os
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ═══════════════════════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════════════════════

def _parse_float(val):
    """Parse a float value, returning 0.0 for empty/invalid.
    Same behavior as shared/loader.py:parse_float."""
    if val is None or val == "" or val == "":
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def _cell_str(val):
    """Get string from a cell value, handling None."""
    if val is None:
        return ""
    return str(val).strip()


def _read_sheet_as_dicts(ws):
    """Read a worksheet into a list of dicts using row 1 as headers."""
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if not rows:
        return []
    headers = [_cell_str(h) for h in rows[0]]
    result = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        d = {}
        for i, h in enumerate(headers):
            if h and i < len(row):
                d[h] = row[i]
        result.append(d)
    return result


# ═══════════════════════════════════════════════════════════════════════════
# READERS — match shared/loader.py data shapes
# ═══════════════════════════════════════════════════════════════════════════

def load_providers_from_excel(wb):
    """Load provider data from the 'Providers' sheet.

    Returns:
        dict: provider_name -> {shift_type, fte, scheduler, annual_weeks,
              annual_weekends, annual_nights, weeks_remaining, weekends_remaining,
              nights_remaining, pct_cooper, pct_inspira_veb, pct_inspira_mhw,
              pct_mannington, pct_virtua, pct_cape, holiday_1, holiday_2, ...}
    """
    ws = wb["Providers"]
    rows = _read_sheet_as_dicts(ws)
    providers = {}

    for row in rows:
        name = _cell_str(row.get("provider_name"))
        if not name:
            continue

        providers[name] = {
            "shift_type":         _cell_str(row.get("shift_type")),
            "fte":                _parse_float(row.get("fte")),
            "scheduler":          _cell_str(row.get("scheduler")),
            "annual_weeks":       _parse_float(row.get("annual_weeks")),
            "annual_weekends":    _parse_float(row.get("annual_weekends")),
            "annual_nights":      _parse_float(row.get("annual_nights")),
            "prior_weeks_worked":    _parse_float(row.get("prior_weeks_worked")),
            "prior_weekends_worked": _parse_float(row.get("prior_weekends_worked")),
            "prior_nights_worked":   _parse_float(row.get("prior_nights_worked")),
            "weeks_remaining":    _parse_float(row.get("weeks_remaining")),
            "weekends_remaining": _parse_float(row.get("weekends_remaining")),
            "nights_remaining":   _parse_float(row.get("nights_remaining")),
            "pct_cooper":         _parse_float(row.get("pct_cooper")),
            "pct_inspira_veb":    _parse_float(row.get("pct_inspira_veb")),
            "pct_inspira_mhw":    _parse_float(row.get("pct_inspira_mhw")),
            "pct_mannington":     _parse_float(row.get("pct_mannington")),
            "pct_virtua":         _parse_float(row.get("pct_virtua")),
            "pct_cape":           _parse_float(row.get("pct_cape")),
            "holiday_1":          _cell_str(row.get("holiday_1")),
            "holiday_2":          _cell_str(row.get("holiday_2")),
        }

    return providers


def load_tags_from_excel(wb):
    """Load provider tags from the 'Provider Tags' sheet.

    Returns:
        dict: provider_name -> [{"tag": str, "rule": str}, ...]
    """
    ws = wb["Provider Tags"]
    rows = _read_sheet_as_dicts(ws)
    tags = defaultdict(list)

    for row in rows:
        name = _cell_str(row.get("provider_name"))
        tag = _cell_str(row.get("tag"))
        rule = _cell_str(row.get("rule"))
        if name and tag:
            tags[name].append({"tag": tag, "rule": rule})

    return dict(tags)


def load_tag_definitions_from_excel(wb):
    """Load tag definitions from the 'Tag Definitions' sheet.

    Returns:
        dict: tag_name -> {engine_status, description, rule_format, example, notes}
    """
    ws = wb["Tag Definitions"]
    rows = _read_sheet_as_dicts(ws)
    defs = {}

    # Column headers in the sheet use display names; map to internal keys
    header_map = {
        "Tag Name": "tag_name",
        "tag_name": "tag_name",
        "Engine Status": "engine_status",
        "engine_status": "engine_status",
        "What It Does": "description",
        "description": "description",
        "Expected Rule Format": "rule_format",
        "rule_format": "rule_format",
        "Example": "example",
        "example_rule": "example",
        "example": "example",
        "Notes": "notes",
        "notes": "notes",
    }

    for row in rows:
        # Normalize keys
        norm = {}
        for k, v in row.items():
            mapped = header_map.get(k, k)
            norm[mapped] = _cell_str(v)

        tag_name = norm.get("tag_name", "")
        if not tag_name:
            continue

        defs[tag_name] = {
            "engine_status": norm.get("engine_status", ""),
            "description":   norm.get("description", ""),
            "rule_format":   norm.get("rule_format", ""),
            "example":       norm.get("example", ""),
            "notes":         norm.get("notes", ""),
        }

    return defs


def load_sites_from_excel(wb):
    """Load site demand from the 'Sites' sheet.

    Returns:
        dict: (site_name, day_type) -> providers_needed
    """
    ws = wb["Sites"]
    rows = _read_sheet_as_dicts(ws)
    sites = {}

    for row in rows:
        site = _cell_str(row.get("site"))
        day_type = _cell_str(row.get("day_type"))
        needed = int(_parse_float(row.get("providers_needed")))
        if site and day_type:
            sites[(site, day_type)] = needed

    return sites


# ═══════════════════════════════════════════════════════════════════════════
# WRITER — Tag Review sheet
# ═══════════════════════════════════════════════════════════════════════════

# Styles
_HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9'),
)
_WRAP = Alignment(wrap_text=True, vertical='top')

_STATUS_STYLES = {
    "ACTIVE":  (PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
                Font(bold=True, color="006100", size=10)),
    "PLANNED": (PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
                Font(bold=True, color="9C6500", size=10)),
    "INFO":    (PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),
                Font(bold=True, color="1F4E79", size=10)),
    "UNKNOWN": (PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
                Font(bold=True, color="9C0006", size=10)),
}

_ROW_FILLS = {
    "UNRESOLVED": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
    "UNKNOWN":    PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid"),
}

_ISSUE_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

_COLUMNS = [
    ("provider_name",          14, "Provider (Original)"),
    ("resolved_name",          14, "Provider (Resolved)"),
    ("name_status",             7, "Name Status"),
    ("tag",                    11, "Tag"),
    ("tag_status",              7, "Tag Status"),
    ("rule",                   25, "Rule"),
    ("engine_interpretation",  30, "Engine Interpretation"),
    ("issues",                 25, "Issues"),
]


def write_tag_review_sheet(wb, results, summary):
    """Write (or overwrite) the 'Tag Review' sheet.

    Args:
        wb: openpyxl Workbook
        results: list of dicts, one per evaluated tag
        summary: dict with count fields for the summary block
    """
    sheet_name = "Tag Review"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(sheet_name)

    # Header row
    for ci, (key, width, label) in enumerate(_COLUMNS, 1):
        c = ws.cell(row=1, column=ci, value=label)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = Alignment(horizontal='center', wrap_text=True)
        c.border = _THIN_BORDER
        ws.column_dimensions[chr(64 + ci)].width = width

    # Data rows
    for ri, rec in enumerate(results, 2):
        name_status = rec.get("name_status", "ok")
        tag_status = rec.get("tag_status", "")
        issues_str = rec.get("issues", "")

        # Determine row fill
        row_fill = None
        if name_status == "UNRESOLVED":
            row_fill = _ROW_FILLS["UNRESOLVED"]
        elif tag_status == "UNKNOWN":
            row_fill = _ROW_FILLS["UNKNOWN"]

        for ci, (key, _, _) in enumerate(_COLUMNS, 1):
            val = rec.get(key, "")
            c = ws.cell(row=ri, column=ci, value=val)
            c.border = _THIN_BORDER
            c.alignment = _WRAP

            if row_fill:
                c.fill = row_fill

        # Status cell coloring (column E = tag_status)
        if tag_status in _STATUS_STYLES:
            fill, font = _STATUS_STYLES[tag_status]
            c = ws.cell(row=ri, column=5)
            c.fill = fill
            c.font = font

        # Issue cell coloring (column H)
        if issues_str:
            c = ws.cell(row=ri, column=8)
            c.fill = _ISSUE_FILL

    last_data_row = len(results) + 1

    # Summary block
    sr = last_data_row + 2
    ws.cell(row=sr, column=1, value="SUMMARY").font = Font(bold=True, size=12)
    summary_lines = [
        ("Total tags evaluated", summary.get("total_tags", 0)),
        ("Providers with tags", summary.get("providers_with_tags", 0)),
        ("Tags recognized", summary.get("tags_recognized", 0)),
        ("Tags unrecognized", summary.get("tags_unrecognized", 0)),
        ("  ACTIVE", summary.get("active_count", 0)),
        ("  PLANNED", summary.get("planned_count", 0)),
        ("  INFO", summary.get("info_count", 0)),
        ("Name resolution issues", summary.get("name_issues", 0)),
        ("Parse warnings", summary.get("parse_warnings", 0)),
        ("Data quality issues", summary.get("data_quality_issues", 0)),
    ]
    for i, (label, val) in enumerate(summary_lines):
        ws.cell(row=sr + 1 + i, column=1, value=label).font = Font(bold=True, size=10)
        ws.cell(row=sr + 1 + i, column=2, value=val)

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:H{last_data_row}"


# ═══════════════════════════════════════════════════════════════════════════
# SHARED WRITER HELPERS
# ═══════════════════════════════════════════════════════════════════════════

def _col_letter(n):
    """Convert 1-based column number to letter(s). 1='A', 26='Z', 27='AA'."""
    result = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        result = chr(65 + remainder) + result
    return result


_MATCH_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_DISC_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_MISSING_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
_RED_FONT = Font(color="9C0006", size=10)
_HIGH_FILL = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
_ELEVATED_FILL = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
_MODERATE_FILL = PatternFill(start_color="CFE2FF", end_color="CFE2FF", fill_type="solid")
_MUST_FILL = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
_SHOULD_FILL = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
_UNAVAIL_FILL = PatternFill(start_color="CFE2FF", end_color="CFE2FF", fill_type="solid")


def _write_header_row(ws, columns):
    """Write a styled header row. columns = [(label, width), ...]"""
    for ci, (label, width) in enumerate(columns, 1):
        c = ws.cell(row=1, column=ci, value=label)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = Alignment(horizontal='center', wrap_text=True)
        c.border = _THIN_BORDER
        ws.column_dimensions[_col_letter(ci)].width = width


def _write_summary_block(ws, start_row, lines):
    """Write a summary block below data. lines = [(label, value), ...]"""
    ws.cell(row=start_row, column=1, value="SUMMARY").font = Font(bold=True, size=12)
    for i, (label, val) in enumerate(lines):
        ws.cell(row=start_row + 1 + i, column=1, value=label).font = Font(bold=True, size=10)
        ws.cell(row=start_row + 1 + i, column=2, value=val)


# ═══════════════════════════════════════════════════════════════════════════
# WRITER — Prior Actuals Review (Task 2)
# ═══════════════════════════════════════════════════════════════════════════

_PA_COLUMNS = [
    ("Provider", 22), ("Computed Weeks", 12), ("Excel Weeks", 11),
    ("Weeks Diff", 10), ("Computed WE", 11), ("Excel WE", 10),
    ("WE Diff", 9), ("Status", 14), ("Detail", 35),
]


def write_prior_actuals_review_sheet(wb, pa_result):
    """Write (or overwrite) the 'Prior Actuals Review' sheet."""
    sheet_name = "Prior Actuals Review"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(sheet_name)
    _write_header_row(ws, _PA_COLUMNS)

    comparisons = pa_result.get("comparisons", [])
    for ri, rec in enumerate(comparisons, 2):
        status = rec.get("status", "")

        # Row fill by status
        if status == "MATCH":
            row_fill = _MATCH_FILL
        elif status == "DISCREPANCY":
            row_fill = _DISC_FILL
        elif status.startswith("MISSING"):
            row_fill = _MISSING_FILL
        else:
            row_fill = None

        vals = [
            rec.get("provider", ""),
            rec.get("computed_weeks", 0),
            rec.get("excel_weeks", 0),
            rec.get("weeks_diff", 0),
            rec.get("computed_weekends", 0),
            rec.get("excel_weekends", 0),
            rec.get("weekends_diff", 0),
            status,
            rec.get("detail", ""),
        ]

        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.border = _THIN_BORDER
            c.alignment = _WRAP
            if row_fill:
                c.fill = row_fill

        # Red font on diff columns if significant
        if rec.get("weeks_diff", 0) >= 0.5:
            ws.cell(row=ri, column=4).font = _RED_FONT
        if rec.get("weekends_diff", 0) >= 0.5:
            ws.cell(row=ri, column=7).font = _RED_FONT

    last_data_row = len(comparisons) + 1
    s = pa_result.get("summary", {})
    _write_summary_block(ws, last_data_row + 2, [
        ("Providers compared", s.get("providers_compared", 0)),
        ("Matching", s.get("providers_matching", 0)),
        ("Discrepancies", s.get("providers_with_discrepancy", 0)),
        ("Missing from schedule", s.get("missing_from_schedule", 0)),
        ("Missing from Excel", s.get("missing_from_excel", 0)),
        ("Files parsed", f"{s.get('files_parsed', 0)}/{s.get('files_expected', 9)}"),
    ])

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:I{last_data_row}"


# ═══════════════════════════════════════════════════════════════════════════
# WRITER — Scheduling Difficulty (Task 3)
# ═══════════════════════════════════════════════════════════════════════════

_DIFF_COLUMNS = [
    ("Provider", 22), ("FTE", 6), ("Shift Type", 10),
    ("Ann Weeks", 9), ("Remaining Wk", 11), ("Density", 8),
    ("Ann WE", 8), ("Remaining WE", 11), ("WE Density", 9),
    ("Risk", 10), ("Eligible Sites", 11), ("Capacity", 10),
    ("Tags", 18), ("Notes", 40),
]

_DIFF_RETRO_COLUMNS = [
    ("Actual B3 Wk", 11), ("Actual B3 WE", 11),
    ("Max Consec", 10), ("Violation?", 10), ("Accuracy", 10),
]

_RISK_FILLS = {
    "HIGH": _HIGH_FILL,
    "ELEVATED": _ELEVATED_FILL,
    "MODERATE": _MODERATE_FILL,
}

_CORRECT_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_UNDER_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_OVER_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

_ACCURACY_FILLS = {
    "CORRECT": _CORRECT_FILL,
    "CLOSE": _CORRECT_FILL,
    "OVER": _OVER_FILL,
    "UNDER": _UNDER_FILL,
    "MISS": _UNDER_FILL,
}


def write_difficulty_sheet(wb, diff_result, retro_records=None, retro_summary=None):
    """Write (or overwrite) the 'Scheduling Difficulty' sheet.

    Args:
        wb: openpyxl Workbook
        diff_result: dict from evaluate_difficulty()
        retro_records: optional list from evaluate_difficulty_retrospective()
        retro_summary: optional dict from evaluate_difficulty_retrospective()
    """
    sheet_name = "Scheduling Difficulty"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    has_retro = retro_records is not None
    columns = _DIFF_COLUMNS + (_DIFF_RETRO_COLUMNS if has_retro else [])

    ws = wb.create_sheet(sheet_name)
    _write_header_row(ws, columns)

    records = retro_records if has_retro else diff_result.get("records", [])
    for ri, rec in enumerate(records, 2):
        risk = rec.get("risk_level", "LOW")
        row_fill = _RISK_FILLS.get(risk)

        vals = [
            rec.get("provider", ""),
            rec.get("fte", 0),
            rec.get("shift_type", ""),
            rec.get("annual_weeks", 0),
            rec.get("remaining_weeks", 0),
            f"{rec.get('density', 0):.0%}",
            rec.get("annual_weekends", 0),
            rec.get("remaining_weekends", 0),
            f"{rec.get('weekend_density', 0):.0%}",
            risk,
            rec.get("num_eligible_sites", 0),
            rec.get("capacity_status", ""),
            ", ".join(rec.get("tag_badges", [])),
            "; ".join(rec.get("notes", [])),
        ]

        if has_retro:
            accuracy = rec.get("prediction_accuracy", "")
            max_c = rec.get("actual_max_consecutive", 0)
            violation = ""
            if rec.get("actual_hard_violation"):
                violation = "HARD (>12)"
            elif rec.get("actual_extended_stretch"):
                violation = "Extended (8-12)"
            vals.extend([
                rec.get("actual_b3_weeks", 0),
                rec.get("actual_b3_weekends", 0),
                max_c,
                violation,
                accuracy,
            ])

        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.border = _THIN_BORDER
            c.alignment = _WRAP
            if ci <= len(_DIFF_COLUMNS) and row_fill:
                c.fill = row_fill

        # Color retrospective accuracy column
        if has_retro:
            acc_col = len(_DIFF_COLUMNS) + 5  # Accuracy column
            accuracy = rec.get("prediction_accuracy", "")
            acc_fill = _ACCURACY_FILLS.get(accuracy)
            if acc_fill:
                ws.cell(row=ri, column=acc_col).fill = acc_fill
            # Red font on violation column if HARD
            if rec.get("actual_hard_violation"):
                ws.cell(row=ri, column=len(_DIFF_COLUMNS) + 4).font = _RED_FONT

    last_data_row = len(records) + 1
    s = diff_result.get("summary", {})
    summary_lines = [
        ("Total eligible providers", s.get("total_eligible", 0)),
        ("HIGH risk", s.get("high_count", 0)),
        ("ELEVATED risk", s.get("elevated_count", 0)),
        ("MODERATE risk", s.get("moderate_count", 0)),
        ("LOW risk", s.get("low_count", 0)),
        ("Mean density", f"{s.get('mean_density', 0):.0%}"),
        ("Tight capacity", s.get("tight_capacity_count", 0)),
        ("Excess capacity", s.get("excess_capacity_count", 0)),
        ("Single-site providers", s.get("single_site_count", 0)),
    ]
    if retro_summary:
        summary_lines.append(("", ""))
        summary_lines.append(("── RETROSPECTIVE ──", ""))
        summary_lines.append(("Providers with B3 data", retro_summary.get("providers_with_data", 0)))
        summary_lines.append(("Correct/Close predictions", retro_summary.get("correct_predictions", 0)))
        summary_lines.append(("Over-predictions", retro_summary.get("over_predictions", 0)))
        summary_lines.append(("Under-predictions", retro_summary.get("under_predictions", 0)))
        summary_lines.append(("Actual hard violations (>12)", retro_summary.get("actual_hard_violations", 0)))
        summary_lines.append(("Actual extended only (8-12)", retro_summary.get("actual_extended_only", 0)))
        summary_lines.append(("Accuracy", f"{retro_summary.get('accuracy_pct', 0)}%"))

    _write_summary_block(ws, last_data_row + 2, summary_lines)

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:{_col_letter(len(columns))}{last_data_row}"


# ═══════════════════════════════════════════════════════════════════════════
# WRITER — Holiday Review (Task 4)
# ═══════════════════════════════════════════════════════════════════════════

_HOL_COLUMNS = [
    ("Provider", 22), ("FTE", 6), ("Site Dir", 7),
    ("Required", 8), ("Worked", 7), ("Owe", 6),
    ("Holidays Worked", 28), ("Preferences", 22),
    ("Mem Day Pref?", 11), ("Mem Day Avail?", 12),
    ("Tier", 12), ("Issues", 35),
]

_HOL_RETRO_COLUMNS = [
    ("Actual Mem Day?", 13), ("Tier Outcome", 14),
]

_TIER_FILLS = {
    "MUST": _MUST_FILL,
    "SHOULD": _SHOULD_FILL,
    "UNAVAILABLE": _UNAVAIL_FILL,
}

_OUTCOME_FILLS = {
    "CORRECT": _CORRECT_FILL,
    "HONORED": _CORRECT_FILL,
    "NOT_SCHEDULED": _OVER_FILL,
    "OVERRIDE": _OVER_FILL,
    "EXTRA": _OVER_FILL,
    "ERROR": _UNDER_FILL,
}


def write_holiday_review_sheet(wb, hol_result, retro_records=None, retro_summary=None):
    """Write (or overwrite) the 'Holiday Review' sheet.

    Args:
        wb: openpyxl Workbook
        hol_result: dict from evaluate_holidays()
        retro_records: optional list from evaluate_holiday_retrospective()
        retro_summary: optional dict from evaluate_holiday_retrospective()
    """
    sheet_name = "Holiday Review"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    has_retro = retro_records is not None
    columns = _HOL_COLUMNS + (_HOL_RETRO_COLUMNS if has_retro else [])

    ws = wb.create_sheet(sheet_name)
    _write_header_row(ws, columns)

    records = retro_records if has_retro else hol_result.get("records", [])
    for ri, rec in enumerate(records, 2):
        tier = rec.get("tier", "")
        row_fill = _TIER_FILLS.get(tier)

        # Build issues string
        issue_parts = []
        if rec.get("worked_both_xmas_ny"):
            issue_parts.append("Worked both Xmas + NY")
        if rec.get("worked_neither_xmas_ny") and rec.get("still_owe", 0) > 0:
            issue_parts.append("Worked neither Xmas nor NY")
        if rec.get("prefs_violated"):
            issue_parts.append(f"Pref violated: {', '.join(rec['prefs_violated'])}")
        if rec.get("still_owe", 0) >= 2:
            issue_parts.append(f"Impossible: owes {rec['still_owe']} with 1 holiday left")
        if rec.get("count_worked", 0) > rec.get("required", 0) and rec.get("required", 0) > 0:
            issue_parts.append(f"Overworked: {rec['count_worked']}/{rec['required']}")

        vals = [
            rec.get("provider", ""),
            rec.get("fte", 0),
            "Yes" if rec.get("is_site_director") else "",
            rec.get("required", 0),
            rec.get("count_worked", 0),
            rec.get("still_owe", 0),
            ", ".join(rec.get("holidays_worked", [])),
            ", ".join(rec.get("preferences", [])),
            "Yes" if rec.get("memorial_is_preference") else "",
            "Yes" if rec.get("mem_available") else "No",
            tier,
            "; ".join(issue_parts),
        ]

        if has_retro:
            actual = rec.get("actual_worked_memorial")
            outcome = rec.get("tier_outcome", "")
            if actual is None:
                vals.append("N/A")
            else:
                vals.append("Yes" if actual else "No")
            vals.append(outcome)

        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.border = _THIN_BORDER
            c.alignment = _WRAP
            if ci <= len(_HOL_COLUMNS) and row_fill:
                c.fill = row_fill

        # Red font on Owe if >= 2 (impossible)
        if rec.get("still_owe", 0) >= 2:
            ws.cell(row=ri, column=6).font = _RED_FONT

        # Red font on availability if No
        if not rec.get("mem_available"):
            ws.cell(row=ri, column=10).font = _RED_FONT

        # Color retrospective outcome column
        if has_retro:
            outcome_col = len(_HOL_COLUMNS) + 2  # Tier Outcome column
            outcome = rec.get("tier_outcome", "")
            outcome_fill = _OUTCOME_FILLS.get(outcome)
            if outcome_fill:
                ws.cell(row=ri, column=outcome_col).fill = outcome_fill
            if outcome == "ERROR":
                ws.cell(row=ri, column=outcome_col).font = _RED_FONT

    last_data_row = len(records) + 1
    s = hol_result.get("summary", {})
    by_tier = s.get("by_tier", {})
    issue_counts = s.get("issue_counts", {})
    summary_lines = [
        ("Total eligible providers", s.get("total_eligible", 0)),
        ("MUST (Memorial Day)", by_tier.get("MUST", 0)),
        ("SHOULD (pref conflict)", by_tier.get("SHOULD", 0)),
        ("MET", by_tier.get("MET", 0)),
        ("UNAVAILABLE", by_tier.get("UNAVAILABLE", 0)),
        ("EXEMPT", by_tier.get("EXEMPT", 0)),
        ("Total issues", s.get("total_issues", 0)),
        ("  Overworked", issue_counts.get("overworked", 0)),
        ("  Both Xmas+NY", issue_counts.get("both_xmas_ny", 0)),
        ("  Neither Xmas nor NY", issue_counts.get("neither_xmas_ny", 0)),
        ("  Pref violated", issue_counts.get("pref_violated", 0)),
        ("  Impossible", issue_counts.get("impossible", 0)),
    ]
    if retro_summary:
        summary_lines.append(("", ""))
        summary_lines.append(("── RETROSPECTIVE ──", ""))
        summary_lines.append(("Providers with B3 data", retro_summary.get("providers_with_data", 0)))
        summary_lines.append(("Actually worked Memorial Day", retro_summary.get("actually_worked_memorial", 0)))
        summary_lines.append(("MUST tier who worked", f"{retro_summary.get('must_worked', 0)}/{retro_summary.get('must_total', 0)} ({retro_summary.get('must_rate', 0)}%)"))
        summary_lines.append(("Unavailable errors", retro_summary.get("unavailable_errors", 0)))
        summary_lines.append(("Preference overrides", retro_summary.get("preference_overrides", 0)))

    _write_summary_block(ws, last_data_row + 2, summary_lines)

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:{_col_letter(len(columns))}{last_data_row}"
