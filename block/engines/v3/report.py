#!/usr/bin/env python3
"""
HTML Report Generator for Block Schedule Engine v3.

Generates a single interactive HTML report with multiple tabs:
  0. Dashboard   — executive summary, stat cards, key metrics
  1. Calendar    — weekly schedule grid by site
  2. Sites       — per-site coverage analysis with weekly detail
  3. Providers   — collapsible provider cards with mini calendars
  4. Utilization — provider utilization table, distribution charts
  5. Stretches   — consecutive-day analysis, stretch distribution
  6. Gaps        — gap report with candidate lists (when gaps exist)
"""

import calendar
import hashlib
import html as html_mod
import json
import os
import sys
from collections import defaultdict
from datetime import datetime, timedelta

_PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", ".."))
if _PROJECT_ROOT not in sys.path:
    sys.path.insert(0, _PROJECT_ROOT)

from block.engines.shared.loader import (
    SITE_COLORS, SITE_SHORT, SITE_PCT_MAP, SCHEDULES_DIR,
    load_availability, get_eligible_sites,
)
from block.engines.v3.excel_io import (
    load_providers_from_excel, load_tags_from_excel, load_sites_from_excel,
)
from name_match import match_provider

try:
    from openpyxl import load_workbook as _load_workbook
except ImportError:
    _load_workbook = None


# ─── Helpers ─────────────────────────────────────────────────────────────────

def esc(s):
    """HTML-escape a string."""
    return html_mod.escape(str(s)) if s is not None else ""


def prov_id(name):
    """Create a safe HTML id from provider name."""
    return "prov-" + name.replace(" ", "-").replace(",", "").replace("'", "").replace(".", "")


def _pct(num, denom):
    """Safe percentage calculation."""
    return (num / denom * 100) if denom > 0 else 0


def _load_password():
    """Load report password from config.json, return SHA-256 hash or empty string."""
    config_path = os.path.join(_PROJECT_ROOT, "config.json")
    if not os.path.exists(config_path):
        return ""
    try:
        with open(config_path) as f:
            config = json.load(f)
        password = config.get("report_password", "")
        if password:
            return hashlib.sha256(password.encode("utf-8")).hexdigest()
    except (json.JSONDecodeError, IOError):
        pass
    return ""


def _load_full_availability():
    """Load all availability JSONs for mini calendar rendering."""
    avail_all = {}
    if not os.path.isdir(SCHEDULES_DIR):
        return avail_all
    for fname in os.listdir(SCHEDULES_DIR):
        if not fname.endswith(".json"):
            continue
        fpath = os.path.join(SCHEDULES_DIR, fname)
        try:
            with open(fpath) as f:
                data = json.load(f)
        except (json.JSONDecodeError, IOError):
            continue
        name = data.get("name", "").strip()
        if not name:
            continue
        if name not in avail_all:
            avail_all[name] = {}
        for day in data.get("days", []):
            avail_all[name][day["date"]] = day.get("status", "blank")
    return avail_all


# ─── CSS ─────────────────────────────────────────────────────────────────────

def _common_css():
    return """
:root {
  --bg: #ffffff; --text: #1a1a1a; --heading: #0d47a1;
  --border: #d0d0d0; --stripe: #f5f7fa; --link: #1565c0;
  --weekend-bg: #e3f2fd; --short-bg: #fce4ec; --short-text: #b71c1c;
  --ok-bg: #e8f5e9; --ok-text: #2e7d32; --flex-bg: #fff3e0;
  --warn-bg: #fff3e0; --warn-text: #e65100;
}
* { box-sizing: border-box; margin: 0; padding: 0; }
body {
  font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
  font-size: 13px; color: var(--text); background: var(--bg); line-height: 1.4;
}
.container { max-width: 1600px; margin: 0 auto; padding: 16px 24px; }
h1 { font-size: 22px; color: var(--heading); margin-bottom: 4px; }
h2 { font-size: 17px; color: var(--heading); margin: 24px 0 8px; border-bottom: 2px solid var(--heading); padding-bottom: 4px; }
h3 { font-size: 14px; color: #333; margin: 16px 0 6px; }
.subtitle { color: #666; font-size: 13px; margin-bottom: 16px; }
.generated { color: #999; font-size: 11px; margin-top: 4px; }

/* Tabs */
.tabs { display: flex; gap: 0; border-bottom: 2px solid var(--heading); margin-bottom: 16px; flex-wrap: wrap; }
.tab { padding: 8px 16px; cursor: pointer; border: 1px solid var(--border); border-bottom: none;
  border-radius: 6px 6px 0 0; background: #f9f9f9; font-size: 13px; font-weight: 500;
  color: #555; margin-bottom: -2px; }
.tab:hover { background: #e3f2fd; }
.tab.active { background: var(--heading); color: white; border-color: var(--heading); }
.tab-content { display: none; }
.tab-content.active { display: block; }

/* Tables */
table { border-collapse: collapse; width: 100%; margin-bottom: 16px; }
th, td { border: 1px solid var(--border); padding: 4px 8px; text-align: left; vertical-align: top; }
th { background: #e8eaf6; font-weight: 600; font-size: 12px; position: sticky; top: 0; z-index: 2;
  cursor: pointer; user-select: none; }
th:hover { background: #c5cae9; }
th.sort-asc::after { content: " ▲"; font-size: 9px; }
th.sort-desc::after { content: " ▼"; font-size: 9px; }
tr:nth-child(even) { background: var(--stripe); }
tr.weekend-row { background: var(--weekend-bg); }
tr.short-row { background: var(--short-bg); }

/* Badges */
.badge { display: inline-block; padding: 1px 6px; border-radius: 3px; font-size: 11px; font-weight: 600; }
.badge-ok { background: var(--ok-bg); color: var(--ok-text); }
.badge-short { background: var(--short-bg); color: var(--short-text); }
.badge-warn { background: var(--warn-bg); color: var(--warn-text); }
.badge-info { background: #e3f2fd; color: #1565c0; }
.site-badge { display: inline-block; padding: 2px 6px; border-radius: 3px; color: white;
  font-size: 11px; font-weight: 600; margin: 1px 2px; white-space: nowrap; }
a.prov-link { color: var(--link); text-decoration: none; cursor: pointer; font-size: 12px;
  display: block; padding: 1px 0; }
a.prov-link:hover { text-decoration: underline; }

/* Stats grid */
.stats-grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(160px, 1fr)); gap: 12px; margin: 16px 0; }
.stat-card { background: #f5f7fa; border: 1px solid var(--border); border-radius: 8px; padding: 12px; text-align: center; }
.stat-value { font-size: 28px; font-weight: 700; color: var(--heading); }
.stat-label { font-size: 11px; color: #666; margin-top: 2px; }
.stat-card.green { border-color: var(--ok-text); }
.stat-card.green .stat-value { color: var(--ok-text); }
.stat-card.red { border-color: var(--short-text); }
.stat-card.red .stat-value { color: var(--short-text); }
.stat-card.orange { border-color: var(--warn-text); }
.stat-card.orange .stat-value { color: var(--warn-text); }

/* Provider detail cards */
.prov-card { border: 1px solid var(--border); border-radius: 8px; margin-bottom: 8px; overflow: hidden; }
.prov-header { padding: 8px 12px; background: #f5f7fa; cursor: pointer; display: flex;
  justify-content: space-between; align-items: center; }
.prov-header:hover { background: #e3f2fd; }
.prov-body { display: none; padding: 12px; border-top: 1px solid var(--border); }
.prov-card.open .prov-body { display: block; }
.prov-arrow { transition: transform 0.2s; }
.prov-card.open .prov-arrow { transform: rotate(90deg); }

/* Utilization bars */
.util-bar { width: 80px; height: 14px; background: #eee; border-radius: 3px; display: inline-block; vertical-align: middle; }
.util-fill { height: 100%; border-radius: 3px; }

/* Mini calendar */
.avail-months { display: flex; gap: 12px; flex-wrap: wrap; margin: 8px 0; }
.avail-month h4 { font-size: 11px; margin-bottom: 4px; text-align: center; }
.mini-cal { display: grid; grid-template-columns: repeat(7, 22px); gap: 1px; }
.dh { font-size: 9px; text-align: center; font-weight: 600; color: #999; }
.dc { width: 22px; height: 22px; font-size: 9px; text-align: center; line-height: 22px;
  border-radius: 2px; position: relative; }
.dc.empty { background: transparent; }
.dc.avail { background: #c8e6c9; }
.dc.unavail { background: #ffcdd2; }
.dc.blank { background: #fff9c4; }
.dn { font-size: 8px; }
.asg { position: absolute; top: 0; left: 0; right: 0; bottom: 0; border-radius: 2px;
  font-size: 7px; color: white; line-height: 22px; text-align: center; opacity: 0.85; }

/* Search */
.search-box { padding: 8px 12px; border: 1px solid var(--border); border-radius: 6px;
  font-size: 13px; width: 300px; margin-bottom: 12px; }

/* Filter buttons */
.filter-btn { padding: 4px 10px; border: 2px solid #ccc; border-radius: 4px; background: white;
  cursor: pointer; font-size: 11px; font-weight: 600; margin: 2px; }
.filter-btn.active { border-color: var(--heading); background: #e3f2fd; }

/* Heatmap */
.hm-scroll { overflow-x: auto; margin: 8px 0; }
.hm-table { border-collapse: collapse; font-size: 11px; }
.hm-table th, .hm-table td { border: 1px solid var(--border); padding: 3px 6px; text-align: center; }
.hm-table th { background: #e8eaf6; font-size: 10px; }
.hm-ok { background: #d4edda; color: #155724; font-weight: 600; }
.hm-warn { background: #fff3cd; color: #856404; font-weight: 600; }
.hm-bad { background: #f8d7da; color: #721c24; font-weight: 700; }
.hm-over { background: #d6eaf8; color: #1a5276; font-weight: 600; }

/* Stretch bars */
.stretch-bar { display: inline-block; height: 16px; border-radius: 2px; vertical-align: middle; margin-right: 4px; }

/* Password gate */
#pw-overlay {
  position: fixed; top: 0; left: 0; right: 0; bottom: 0;
  background: #f5f7fa; z-index: 10000;
  display: flex; align-items: center; justify-content: center;
}
#pw-overlay.hidden { display: none; }
#pw-box {
  background: #fff; border: 1px solid var(--border); border-radius: 12px;
  padding: 40px; max-width: 380px; width: 90%; text-align: center;
  box-shadow: 0 4px 24px rgba(0,0,0,0.1);
}
#pw-box h2 { margin-bottom: 8px; font-size: 20px; color: var(--heading); border: none; }
#pw-box p { color: #666; font-size: 13px; margin-bottom: 20px; }
#pw-input {
  width: 100%; padding: 10px 14px; font-size: 15px;
  border: 2px solid var(--border); border-radius: 6px; outline: none; margin-bottom: 12px;
}
#pw-input:focus { border-color: var(--link); }
#pw-btn {
  width: 100%; padding: 10px; font-size: 15px; font-weight: 600;
  background: var(--heading); color: #fff; border: none; border-radius: 6px; cursor: pointer;
}
#pw-btn:hover { background: #1565c0; }
#pw-error { color: var(--short-text); font-size: 13px; margin-top: 10px; display: none; }
#report-content { display: none; }
#report-content.unlocked { display: block; }

/* Navigation header */
.nav { background: #e8eaf6; padding: 10px 20px; margin-bottom: 20px; border-radius: 6px; display: flex; gap: 8px; align-items: center; flex-wrap: wrap; }
.nav a { font-weight: 500; font-size: 13px; color: var(--link); text-decoration: none; padding: 4px 12px; border-radius: 4px; }
.nav a:hover { background: #c5cae9; }
.nav a.active { background: var(--heading); color: #fff; font-weight: 700; }
.nav .nav-label { font-size: 12px; color: #666; font-weight: 600; margin-right: 4px; }

@media (max-width: 768px) {
  body { padding: 10px; font-size: 12px; }
  .container { padding: 10px; }
  h1 { font-size: 18px; }
  .stats-grid { grid-template-columns: repeat(2, 1fr); gap: 8px; }
  .tab { padding: 6px 10px; font-size: 11px; }
  .search-box { width: 100%; }
  .mini-cal { grid-template-columns: repeat(7, 18px); }
  .dc { width: 18px; height: 18px; line-height: 18px; }
}
"""


# ─── Mini Calendar Renderer ─────────────────────────────────────────────────

def _render_mini_calendar(avail_map, date_assignments, block_start, block_end):
    """Render a compact 4-month availability calendar with site assignments."""
    block_months = [(2026, 3), (2026, 4), (2026, 5), (2026, 6)]
    month_names = {3: "March", 4: "April", 5: "May", 6: "June"}
    day_headers = ["S", "M", "T", "W", "T", "F", "S"]

    h = ['<div class="avail-months">']
    for year, month in block_months:
        h.append(f'<div class="avail-month"><h4>{month_names[month]} {year}</h4>')
        h.append('<div class="mini-cal">')
        for dh in day_headers:
            h.append(f'<div class="dh">{dh}</div>')

        first_dow = calendar.weekday(year, month, 1)
        first_dow_sun = (first_dow + 1) % 7
        days_in_month = calendar.monthrange(year, month)[1]

        for _ in range(first_dow_sun):
            h.append('<div class="dc empty"></div>')

        for day in range(1, days_in_month + 1):
            date_str = f"{year}-{month:02d}-{day:02d}"
            status = avail_map.get(date_str, "blank")
            cls = "dc"
            if status == "available":
                cls += " avail"
            elif status == "unavailable":
                cls += " unavail"
            else:
                cls += " blank"

            dt = datetime(year, month, day)
            in_block = block_start <= dt <= block_end

            h.append(f'<div class="{cls}"><div class="dn">{day}</div>')
            assigned_site = date_assignments.get(date_str)
            if assigned_site and in_block:
                color = SITE_COLORS.get(assigned_site, "#666")
                short = SITE_SHORT.get(assigned_site, assigned_site[:3])
                h.append(f'<div class="asg" style="background:{color}">{esc(short)}</div>')
            h.append('</div>')

        total_cells = first_dow_sun + days_in_month
        remaining = (7 - (total_cells % 7)) % 7
        for _ in range(remaining):
            h.append('<div class="dc empty"></div>')
        h.append('</div></div>')

    h.append('</div>')
    return "\n".join(h)


# ═════════════════════════════════════════════════════════════════════════════
# REPORT GENERATION
# ═════════════════════════════════════════════════════════════════════════════

def generate_report(results, output_dir, filename="v3_report.html", nav_html=""):
    """Generate a single comprehensive HTML report for V3 engine output.

    Args:
        results: engine output dict with stats, draft_schedule, gap_report,
                 provider_summary, site_coverage, periods
        filename: output filename
        nav_html: optional HTML for cross-seed navigation header
    """
    stats = results["stats"]
    draft_schedule = results["draft_schedule"]
    gap_report = results["gap_report"]
    provider_summary = results["provider_summary"]
    site_coverage = results["site_coverage"]
    periods = results["periods"]
    seed = stats["seed"]

    all_sites = sorted(site_coverage.keys())

    # Derive block start/end from periods
    all_dates = []
    for p in periods:
        all_dates.extend(p["dates"])
    block_start = datetime.strptime(min(all_dates), "%Y-%m-%d")
    block_end = datetime.strptime(max(all_dates), "%Y-%m-%d")
    date_range = f"{block_start.strftime('%B %d')} &ndash; {block_end.strftime('%B %d, %Y')}"

    # Build per-provider date->site assignments for mini calendars
    prov_date_asgn = defaultdict(dict)
    for period in draft_schedule:
        pidx = period["period_idx"]
        p = periods[pidx]
        for asgn in period["assignments"]:
            pname = asgn["provider"]
            site = asgn["site"]
            for d in p["dates"]:
                prov_date_asgn[pname][d] = site

    # Build per-site weekly fill data
    site_weekly = defaultdict(lambda: defaultdict(dict))
    for period in draft_schedule:
        pidx = period["period_idx"]
        p = periods[pidx]
        ptype = p["type"]
        wnum = p["num"]
        # Count assignments per site for this period
        site_counts = defaultdict(int)
        for asgn in period["assignments"]:
            site_counts[asgn["site"]] += 1
        for site in all_sites:
            site_weekly[site][(wnum, ptype)] = site_counts.get(site, 0)

    # Load availability for mini calendars
    avail_all = _load_full_availability()
    json_names = list(avail_all.keys())

    sorted_provs = sorted(provider_summary.keys())
    now = datetime.now().strftime("%Y-%m-%d %H:%M")

    # Password gate
    pw_hash = _load_password()

    # ─── Compute stretch distribution ────────────────────────────────────
    stretch_buckets = {"0": 0, "1-5": 0, "6-7": 0, "8-9": 0, "10-12": 0, "13+": 0}
    for ps in provider_summary.values():
        mc = ps["max_consecutive_days"]
        if mc == 0:
            stretch_buckets["0"] += 1
        elif mc <= 5:
            stretch_buckets["1-5"] += 1
        elif mc <= 7:
            stretch_buckets["6-7"] += 1
        elif mc <= 9:
            stretch_buckets["8-9"] += 1
        elif mc <= 12:
            stretch_buckets["10-12"] += 1
        else:
            stretch_buckets["13+"] += 1

    # Total demand
    total_wk_demand = sum(sc["weekday_demand"] for sc in site_coverage.values())
    total_we_demand = sum(sc["weekend_demand"] for sc in site_coverage.values())
    total_wk_filled = sum(sc["weekday_filled"] for sc in site_coverage.values())
    total_we_filled = sum(sc["weekend_filled"] for sc in site_coverage.values())

    # ═══ Begin HTML ══════════════════════════════════════════════════════
    pw_gate_html = ""
    if pw_hash:
        pw_gate_html = ('<div id="pw-overlay"><div id="pw-box">'
                        '<h2>Password Required</h2>'
                        '<p>This report contains protected information.</p>'
                        '<input type="password" id="pw-input" placeholder="Enter password" autocomplete="off">'
                        '<button id="pw-btn">Unlock</button>'
                        '<div id="pw-error">Incorrect password</div>'
                        '</div></div>')

    content_cls = "unlocked" if not pw_hash else ""

    h = [f"""<!DOCTYPE html>
<html lang="en"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Block 3 Schedule — V3 Engine (seed={seed})</title>
<style>{_common_css()}</style></head><body>

{pw_gate_html}

<div id="report-content" class="{content_cls}">
<div class="container">

{nav_html}

<h1>Block 3 Schedule Report — V3 Engine</h1>
<p class="subtitle">{date_range} &nbsp;|&nbsp; Seed: {seed}</p>
<p class="generated">Generated: {now}</p>

<div class="tabs">
  <div class="tab active" onclick="switchTab(0)">Dashboard</div>
  <div class="tab" onclick="switchTab(1)">Calendar</div>
  <div class="tab" onclick="switchTab(2)">Sites</div>
  <div class="tab" onclick="switchTab(3)">Providers</div>
  <div class="tab" onclick="switchTab(4)">Utilization</div>
  <div class="tab" onclick="switchTab(5)">Stretches</div>
  <div class="tab" onclick="switchTab(6)">Gaps</div>
</div>
"""]

    # ═══ TAB 0: DASHBOARD ════════════════════════════════════════════════
    h.append('<div class="tab-content active" id="tab0">')
    h.append('<h2>Executive Summary</h2>')

    # Main stat cards
    gap_cls = "green" if stats["total_gaps"] == 0 else "red"
    viol_cls = "green" if stats["zero_gap_violations"] == 0 else "red"
    wk_cls = "green" if stats["weekday_coverage_pct"] == 100 else ("orange" if stats["weekday_coverage_pct"] >= 95 else "red")
    we_cls = "green" if stats["weekend_coverage_pct"] == 100 else ("orange" if stats["weekend_coverage_pct"] >= 95 else "red")

    h.append(f"""<div class="stats-grid">
<div class="stat-card {wk_cls}"><div class="stat-value">{stats['weekday_coverage_pct']:.1f}%</div><div class="stat-label">Weekday Coverage</div></div>
<div class="stat-card {we_cls}"><div class="stat-value">{stats['weekend_coverage_pct']:.1f}%</div><div class="stat-label">Weekend Coverage</div></div>
<div class="stat-card {gap_cls}"><div class="stat-value">{stats['total_gaps']}</div><div class="stat-label">Total Gaps</div></div>
<div class="stat-card {viol_cls}"><div class="stat-value">{stats['zero_gap_violations']}</div><div class="stat-label">Zero-Gap Violations</div></div>
<div class="stat-card"><div class="stat-value">{stats['total_eligible']}</div><div class="stat-label">Eligible Providers</div></div>
<div class="stat-card"><div class="stat-value">{stats['total_weeks_assigned']}</div><div class="stat-label">Weeks Assigned</div></div>
<div class="stat-card"><div class="stat-value">{stats['total_weekends_assigned']}</div><div class="stat-label">Weekends Assigned</div></div>
<div class="stat-card"><div class="stat-value">{stretch_buckets.get('13+', 0)}</div><div class="stat-label">Hard Violations (&gt;12d)</div></div>
</div>""")

    # Site coverage summary table
    h.append('<h2>Site Coverage Summary</h2>')
    h.append('<table><thead><tr><th>Site</th><th>Gap Tolerance</th>'
             '<th>WK Demand</th><th>WK Filled</th><th>WK %</th>'
             '<th>WE Demand</th><th>WE Filled</th><th>WE %</th></tr></thead><tbody>')

    for site in all_sites:
        sc = site_coverage[site]
        color = SITE_COLORS.get(site, "#666")
        wk_pct = sc["weekday_coverage_pct"]
        we_pct = sc["weekend_coverage_pct"]
        wk_badge = f'<span class="badge badge-ok">{wk_pct:.0f}%</span>' if wk_pct == 100 else f'<span class="badge badge-short">{wk_pct:.1f}%</span>'
        we_badge = f'<span class="badge badge-ok">{we_pct:.0f}%</span>' if we_pct == 100 else f'<span class="badge badge-short">{we_pct:.1f}%</span>'
        tol_label = "Zero-gap" if sc["gap_tolerance"] == 0 else str(sc["gap_tolerance"])

        h.append(f'<tr><td><span class="site-badge" style="background:{color}">{esc(site)}</span></td>')
        h.append(f'<td>{tol_label}</td>')
        h.append(f'<td>{sc["weekday_demand"]}</td><td>{sc["weekday_filled"]}</td><td>{wk_badge}</td>')
        h.append(f'<td>{sc["weekend_demand"]}</td><td>{sc["weekend_filled"]}</td><td>{we_badge}</td></tr>')

    h.append(f'<tr style="font-weight:700;background:#e8eaf6"><td>Total</td><td></td>')
    h.append(f'<td>{total_wk_demand}</td><td>{total_wk_filled}</td><td>{_pct(total_wk_filled, total_wk_demand):.1f}%</td>')
    h.append(f'<td>{total_we_demand}</td><td>{total_we_filled}</td><td>{_pct(total_we_filled, total_we_demand):.1f}%</td></tr>')
    h.append('</tbody></table>')

    # Stretch distribution summary
    h.append('<h2>Stretch Distribution</h2>')
    h.append('<table><thead><tr><th>Max Consecutive Days</th><th>Providers</th><th>Assessment</th></tr></thead><tbody>')
    for label, count in stretch_buckets.items():
        if label == "0":
            assess = '<span class="badge badge-info">Unassigned</span>'
        elif label in ("1-5", "6-7"):
            assess = '<span class="badge badge-ok">Normal</span>'
        elif label in ("8-9", "10-12"):
            assess = '<span class="badge badge-warn">Extended</span>'
        else:
            assess = '<span class="badge badge-short">VIOLATION</span>'
        h.append(f'<tr><td>{label}</td><td>{count}</td><td>{assess}</td></tr>')
    h.append('</tbody></table>')

    # Provider utilization summary
    at_cap = stats["providers_at_capacity"]
    under = stats["providers_under_utilized"]

    # Compute totals across all providers
    total_annual_wk = sum(ps["annual_weeks"] for ps in provider_summary.values())
    total_annual_we = sum(ps["annual_weekends"] for ps in provider_summary.values())
    total_prior_wk = sum(ps.get("weeks_prior", 0) for ps in provider_summary.values())
    total_prior_we = sum(ps.get("weekends_prior", 0) for ps in provider_summary.values())
    total_target_wk = sum(ps["weeks_target"] for ps in provider_summary.values())
    total_target_we = sum(ps["weekends_target"] for ps in provider_summary.values())
    total_still_owed_wk = sum(max(0, ps["weeks_target"] - ps["weeks_assigned"]) for ps in provider_summary.values())
    total_still_owed_we = sum(max(0, ps["weekends_target"] - ps["weekends_assigned"]) for ps in provider_summary.values())

    h.append('<h2>Provider Utilization — Annual Picture</h2>')
    h.append(f"""<div class="stats-grid">
<div class="stat-card"><div class="stat-value">{total_annual_wk:.0f}</div><div class="stat-label">Annual WK Obligation</div></div>
<div class="stat-card"><div class="stat-value">{total_prior_wk:.0f}</div><div class="stat-label">B1+B2 WK Worked</div></div>
<div class="stat-card"><div class="stat-value">{total_target_wk}</div><div class="stat-label">B3 WK Owed (remaining)</div></div>
<div class="stat-card green"><div class="stat-value">{stats['total_weeks_assigned']}</div><div class="stat-label">B3 WK Assigned</div></div>
<div class="stat-card"><div class="stat-value">{total_annual_we:.0f}</div><div class="stat-label">Annual WE Obligation</div></div>
<div class="stat-card"><div class="stat-value">{total_prior_we:.0f}</div><div class="stat-label">B1+B2 WE Worked</div></div>
<div class="stat-card"><div class="stat-value">{total_target_we}</div><div class="stat-label">B3 WE Owed (remaining)</div></div>
<div class="stat-card green"><div class="stat-value">{stats['total_weekends_assigned']}</div><div class="stat-label">B3 WE Assigned</div></div>
</div>""")
    h.append(f"""<div class="stats-grid">
<div class="stat-card green"><div class="stat-value">{at_cap}</div><div class="stat-label">Fully Scheduled</div></div>
<div class="stat-card {"orange" if under > 0 else "green"}"><div class="stat-value">{under}</div><div class="stat-label">Still Owe After B3</div></div>
<div class="stat-card {"orange" if total_still_owed_wk > 0 else "green"}"><div class="stat-value">{total_still_owed_wk}</div><div class="stat-label">Unscheduled WK</div></div>
<div class="stat-card {"orange" if total_still_owed_we > 0 else "green"}"><div class="stat-value">{total_still_owed_we}</div><div class="stat-label">Unscheduled WE</div></div>
</div>""")
    h.append('<p style="color:#666;font-size:12px;margin-top:-8px">'
             'B3 target = what each provider still owes for the year (annual - B1 - B2). '
             'Providers whose remaining exceeds what 17 block weeks can absorb will still have unscheduled weeks.</p>')

    h.append('</div>')  # end tab0

    # ═══ TAB 1: CALENDAR ═════════════════════════════════════════════════
    h.append('<div class="tab-content" id="tab1">')
    h.append('<h2>Weekly Schedule Calendar</h2>')

    # Site filter buttons
    h.append('<div style="margin-bottom:12px">')
    h.append('<button class="filter-btn active" onclick="filterSite(\'all\',this)">All Sites</button>')
    for site in all_sites:
        color = SITE_COLORS.get(site, "#666")
        short = SITE_SHORT.get(site, site[:3])
        h.append(f'<button class="filter-btn" onclick="filterSite(\'{esc(site)}\',this)" '
                 f'style="border-color:{color};color:{color}">{esc(short)}</button>')
    h.append('</div>')

    # Weekly schedule table
    for period in draft_schedule:
        pidx = period["period_idx"]
        p = periods[pidx]
        ptype = p["type"]
        wnum = p["num"]
        is_weekend = ptype == "weekend"
        row_cls = ' class="weekend-row"' if is_weekend else ""
        type_label = "WE" if is_weekend else "WK"
        dates_str = f"{p['dates'][0]} to {p['dates'][-1]}"

        h.append(f'<h3>Week {wnum} {type_label}: {dates_str}</h3>')

        # Group assignments by site
        by_site = defaultdict(list)
        for asgn in period["assignments"]:
            by_site[asgn["site"]].append(asgn["provider"])

        if not by_site:
            h.append('<p style="color:#999">No assignments</p>')
            continue

        h.append(f'<div style="display:flex;flex-wrap:wrap;gap:12px;margin-bottom:16px">')
        for site in all_sites:
            provs = by_site.get(site, [])
            if not provs:
                continue
            color = SITE_COLORS.get(site, "#666")
            short = SITE_SHORT.get(site, site[:3])

            # Get demand for this site/type
            sc = site_coverage[site]
            demand_key = "weekend_demand" if is_weekend else "weekday_demand"
            demand_per_week = sc[demand_key] // 17 if sc[demand_key] > 0 else 0

            fill_cls = "badge-ok" if len(provs) >= demand_per_week else "badge-short"

            h.append(f'<div class="site-col" data-site="{esc(site)}" '
                     f'style="border:1px solid {color};border-radius:6px;padding:8px;min-width:160px">')
            h.append(f'<div style="font-weight:600;color:{color};margin-bottom:4px">'
                     f'{esc(site)} <span class="badge {fill_cls}">{len(provs)}/{demand_per_week}</span></div>')
            for prov in sorted(provs):
                pid = prov_id(prov)
                h.append(f'<a class="prov-link" onclick="showProvider(\'{pid}\')">{esc(prov)}</a>')
            h.append('</div>')

        h.append('</div>')

    h.append('</div>')  # end tab1

    # ═══ TAB 2: SITES ════════════════════════════════════════════════════
    h.append('<div class="tab-content" id="tab2">')
    h.append('<h2>Site Coverage Analysis</h2>')

    # Heatmap: sites × weeks
    h.append('<h3>Weekly Fill Heatmap</h3>')
    h.append('<div class="hm-scroll"><table class="hm-table"><thead><tr><th>Site</th>')
    week_nums = sorted(set(p["num"] for p in periods))
    for wn in week_nums:
        h.append(f'<th colspan="2">W{wn}</th>')
    h.append('</tr><tr><th></th>')
    for wn in week_nums:
        h.append('<th>WK</th><th>WE</th>')
    h.append('</tr></thead><tbody>')

    for site in all_sites:
        color = SITE_COLORS.get(site, "#666")
        sc = site_coverage[site]
        wk_per = sc["weekday_demand"] // 17 if sc["weekday_demand"] > 0 else 0
        we_per = sc["weekend_demand"] // 17 if sc["weekend_demand"] > 0 else 0

        h.append(f'<tr><td style="text-align:left;font-weight:600">'
                 f'<span class="site-badge" style="background:{color}">{esc(SITE_SHORT.get(site, site[:3]))}</span> '
                 f'{esc(site)}</td>')

        for wn in week_nums:
            wk_fill = site_weekly[site].get((wn, "week"), 0)
            we_fill = site_weekly[site].get((wn, "weekend"), 0)

            # Weekday cell
            if wk_per == 0:
                h.append('<td>—</td>')
            else:
                diff = wk_fill - wk_per
                if diff == 0:
                    cls = "hm-ok"
                elif diff < 0:
                    cls = "hm-bad" if diff <= -2 else "hm-warn"
                else:
                    cls = "hm-over"
                h.append(f'<td class="{cls}">{wk_fill}</td>')

            # Weekend cell
            if we_per == 0:
                h.append('<td>—</td>')
            else:
                diff = we_fill - we_per
                if diff == 0:
                    cls = "hm-ok"
                elif diff < 0:
                    cls = "hm-bad" if diff <= -2 else "hm-warn"
                else:
                    cls = "hm-over"
                h.append(f'<td class="{cls}">{we_fill}</td>')

        h.append('</tr>')

    h.append('</tbody></table></div>')

    # Per-site detail
    for site in all_sites:
        sc = site_coverage[site]
        color = SITE_COLORS.get(site, "#666")
        wk_per = sc["weekday_demand"] // 17 if sc["weekday_demand"] > 0 else 0
        we_per = sc["weekend_demand"] // 17 if sc["weekend_demand"] > 0 else 0

        h.append(f'<h3><span class="site-badge" style="background:{color}">{esc(site)}</span> '
                 f'Demand: {wk_per}/wk, {we_per}/we &nbsp; Coverage: '
                 f'{sc["weekday_coverage_pct"]:.0f}% wk, {sc["weekend_coverage_pct"]:.0f}% we</h3>')

        h.append('<table><thead><tr><th>Week</th><th>Weekday Fill</th><th>Weekend Fill</th></tr></thead><tbody>')
        for wn in week_nums:
            wk_fill = site_weekly[site].get((wn, "week"), 0)
            we_fill = site_weekly[site].get((wn, "weekend"), 0)

            wk_badge_cls = "badge-ok" if wk_fill >= wk_per else "badge-short"
            we_badge_cls = "badge-ok" if we_fill >= we_per else "badge-short"

            h.append(f'<tr><td>Week {wn}</td>')
            h.append(f'<td><span class="badge {wk_badge_cls}">{wk_fill}/{wk_per}</span></td>')
            h.append(f'<td><span class="badge {we_badge_cls}">{we_fill}/{we_per}</span></td></tr>')
        h.append('</tbody></table>')

    h.append('</div>')  # end tab2

    # ═══ TAB 3: PROVIDERS ════════════════════════════════════════════════
    h.append('<div class="tab-content" id="tab3">')
    h.append('<h2>Provider Details</h2>')

    h.append('<input type="text" class="search-box" id="provSearch" '
             'oninput="filterProviders()" placeholder="Search providers...">')

    h.append(f'<p style="color:#666;margin-bottom:12px">{len(sorted_provs)} providers</p>')

    for pname in sorted_provs:
        ps = provider_summary[pname]
        pid = prov_id(pname)
        total_assigned = ps["weeks_assigned"] + ps["weekends_assigned"]
        total_target = ps["weeks_target"] + ps["weekends_target"]
        total_gap = ps["weeks_gap"] + ps["weekends_gap"]

        # Status badge
        if total_gap == 0:
            status = '<span class="badge badge-ok">FULL</span>'
        elif total_gap <= 2:
            status = f'<span class="badge badge-warn">-{total_gap}</span>'
        else:
            status = f'<span class="badge badge-short">-{total_gap}</span>'

        # Site badges
        site_badges = ""
        for s, cnt in sorted(ps["site_distribution"].items()):
            scolor = SITE_COLORS.get(s, "#666")
            sshort = SITE_SHORT.get(s, s[:3])
            site_badges += f'<span class="site-badge" style="background:{scolor}">{sshort}:{cnt}</span> '

        # FTE / shift type
        info = f'{ps["shift_type"]} | FTE {ps["fte"]:.2f}'

        h.append(f'<div class="prov-card" id="{pid}">')
        h.append(f'<div class="prov-header" onclick="toggleCard(this)">')
        h.append(f'<div><strong>{esc(pname)}</strong> &nbsp; {status} &nbsp; '
                 f'<span style="color:#666;font-size:11px">{info}</span> &nbsp; {site_badges}</div>')
        h.append(f'<div class="prov-arrow">&#9654;</div></div>')

        h.append(f'<div class="prov-body">')

        # Summary stats — full annual picture
        wk_prior = ps.get("weeks_prior", 0)
        we_prior = ps.get("weekends_prior", 0)
        wk_after = max(0, ps["weeks_target"] - ps["weeks_assigned"])
        we_after = max(0, ps["weekends_target"] - ps["weekends_assigned"])
        fs_wk = ps.get("fair_share_wk", "—")
        fs_we = ps.get("fair_share_we", "—")

        h.append('<table style="width:auto;margin-bottom:8px"><thead><tr>'
                 '<th></th><th>Annual</th><th>B1+B2</th><th>B3 Owed</th>'
                 '<th>B3 Assigned</th><th>Still Owed</th><th>Fair Share</th></tr></thead><tbody>')
        h.append(f'<tr><td><strong>Weeks</strong></td>'
                 f'<td>{ps["annual_weeks"]}</td><td>{wk_prior}</td><td>{ps["weeks_target"]}</td>'
                 f'<td>{ps["weeks_assigned"]}</td><td>{wk_after}</td><td>{fs_wk}</td></tr>')
        h.append(f'<tr><td><strong>Weekends</strong></td>'
                 f'<td>{ps["annual_weekends"]}</td><td>{we_prior}</td><td>{ps["weekends_target"]}</td>'
                 f'<td>{ps["weekends_assigned"]}</td><td>{we_after}</td><td>{fs_we}</td></tr>'
                 f'<td>{ps["annual_weekends"]}</td><td>{ps["weekends_remaining"]}</td></tr>')
        h.append('</tbody></table>')

        h.append(f'<p style="margin:4px 0;font-size:12px">Max consecutive: '
                 f'<strong>{ps["max_consecutive_days"]}</strong> days '
                 f'| Eligible sites: {", ".join(ps["eligible_sites"])}</p>')

        # Mini calendar
        json_name = match_provider(pname, json_names)
        avail_map = avail_all.get(json_name, {}) if json_name else {}
        date_asgn = prov_date_asgn.get(pname, {})
        h.append(_render_mini_calendar(avail_map, date_asgn, block_start, block_end))

        # Assignment table
        assignments = ps.get("assignments", [])
        if assignments:
            h.append('<table style="margin-top:8px;width:auto"><thead><tr>'
                     '<th>Wk</th><th>Type</th><th>Dates</th><th>Site</th></tr></thead><tbody>')
            for pidx, site in sorted(assignments, key=lambda x: x[0]):
                if pidx < len(periods):
                    per = periods[pidx]
                    scolor = SITE_COLORS.get(site, "#666")
                    sshort = SITE_SHORT.get(site, site[:3])
                    h.append(f'<tr><td>{per["num"]}</td><td>{per["type"]}</td>'
                             f'<td>{per["dates"][0]} to {per["dates"][-1]}</td>'
                             f'<td><span class="site-badge" style="background:{scolor}">{esc(sshort)}</span></td></tr>')
            h.append('</tbody></table>')

        h.append('</div></div>')  # end prov-body, prov-card

    h.append('</div>')  # end tab3

    # ═══ TAB 4: UTILIZATION ══════════════════════════════════════════════
    h.append('<div class="tab-content" id="tab4">')
    h.append('<h2>Provider Utilization</h2>')

    h.append('<table id="utilTable"><thead><tr>'
             '<th onclick="sortUtil(0,\'str\',this)">Provider</th>'
             '<th onclick="sortUtil(1,\'str\',this)">Type</th>'
             '<th onclick="sortUtil(2,\'num\',this)">FTE</th>'
             '<th onclick="sortUtil(3,\'num\',this)">Annual WK</th>'
             '<th onclick="sortUtil(4,\'num\',this)">B1+B2 WK</th>'
             '<th onclick="sortUtil(5,\'num\',this)">B3 Owed</th>'
             '<th onclick="sortUtil(6,\'num\',this)">B3 Asgn</th>'
             '<th onclick="sortUtil(7,\'num\',this)">WK %</th>'
             '<th onclick="sortUtil(8,\'num\',this)">Annual WE</th>'
             '<th onclick="sortUtil(9,\'num\',this)">B1+B2 WE</th>'
             '<th onclick="sortUtil(10,\'num\',this)">B3 Owed</th>'
             '<th onclick="sortUtil(11,\'num\',this)">B3 Asgn</th>'
             '<th onclick="sortUtil(12,\'num\',this)">WE %</th>'
             '<th onclick="sortUtil(13,\'num\',this)">Consec</th>'
             '<th onclick="sortUtil(14,\'str\',this)">Status</th>'
             '</tr></thead><tbody>')

    for pname in sorted_provs:
        ps = provider_summary[pname]
        wk_prior = ps.get("weeks_prior", 0)
        we_prior = ps.get("weekends_prior", 0)
        wk_pct = _pct(ps["weeks_assigned"], ps["weeks_target"])
        we_pct = _pct(ps["weekends_assigned"], ps["weekends_target"])

        if ps["weeks_gap"] == 0 and ps["weekends_gap"] == 0:
            status = '<span class="badge badge-ok">FULL</span>'
        elif ps["weeks_assigned"] > ps["weeks_target"] or ps["weekends_assigned"] > ps["weekends_target"]:
            status = '<span class="badge badge-warn">OVER</span>'
        else:
            status = '<span class="badge badge-short">OWED</span>'

        # Utilization bars
        wk_bar_pct = min(wk_pct, 100)
        we_bar_pct = min(we_pct, 100)
        wk_color = "#2e7d32" if wk_pct >= 100 else ("#e65100" if wk_pct >= 50 else "#b71c1c")
        we_color = "#2e7d32" if we_pct >= 100 else ("#e65100" if we_pct >= 50 else "#b71c1c")

        h.append(f'<tr><td>{esc(pname)}</td><td>{esc(ps["shift_type"])}</td><td>{ps["fte"]:.2f}</td>')
        h.append(f'<td>{ps["annual_weeks"]}</td><td>{wk_prior}</td><td>{ps["weeks_target"]}</td>')
        h.append(f'<td>{ps["weeks_assigned"]}</td>')
        h.append(f'<td>{wk_pct:.0f}% <div class="util-bar"><div class="util-fill" style="width:{wk_bar_pct}%;background:{wk_color}"></div></div></td>')
        h.append(f'<td>{ps["annual_weekends"]}</td><td>{we_prior}</td><td>{ps["weekends_target"]}</td>')
        h.append(f'<td>{ps["weekends_assigned"]}</td>')
        h.append(f'<td>{we_pct:.0f}% <div class="util-bar"><div class="util-fill" style="width:{we_bar_pct}%;background:{we_color}"></div></div></td>')
        h.append(f'<td>{ps["max_consecutive_days"]}</td>')
        h.append(f'<td>{status}</td></tr>')

    h.append('</tbody></table>')

    # Site distribution summary
    h.append('<h2>Site Assignment Distribution</h2>')
    h.append('<table><thead><tr><th>Site</th><th>Total Assignments</th>'
             '<th>Unique Providers</th><th>Avg per Provider</th></tr></thead><tbody>')

    for site in all_sites:
        total_asgn = 0
        unique_provs = set()
        for pname, ps in provider_summary.items():
            cnt = ps["site_distribution"].get(site, 0)
            if cnt > 0:
                total_asgn += cnt
                unique_provs.add(pname)
        avg = total_asgn / len(unique_provs) if unique_provs else 0
        color = SITE_COLORS.get(site, "#666")
        h.append(f'<tr><td><span class="site-badge" style="background:{color}">{esc(site)}</span></td>')
        h.append(f'<td>{total_asgn}</td><td>{len(unique_provs)}</td><td>{avg:.1f}</td></tr>')

    h.append('</tbody></table>')
    h.append('</div>')  # end tab4

    # ═══ TAB 5: STRETCHES ════════════════════════════════════════════════
    h.append('<div class="tab-content" id="tab5">')
    h.append('<h2>Consecutive Day Analysis</h2>')

    # Stretch distribution visual
    h.append('<div class="stats-grid">')
    stretch_colors = {"0": "#999", "1-5": "#2e7d32", "6-7": "#2e7d32",
                      "8-9": "#e65100", "10-12": "#e65100", "13+": "#b71c1c"}
    for label, count in stretch_buckets.items():
        border = stretch_colors[label]
        h.append(f'<div class="stat-card" style="border-color:{border}">'
                 f'<div class="stat-value" style="color:{border}">{count}</div>'
                 f'<div class="stat-label">{label} days</div></div>')
    h.append('</div>')

    h.append('<p style="color:#666;margin-bottom:12px">'
             'Hard limit: 12 consecutive days. Extended stretches (8-12) are common and acceptable.</p>')

    # Detailed stretch table
    h.append('<h3>Providers by Max Consecutive Days (sorted)</h3>')
    h.append('<table id="stretchTable"><thead><tr>'
             '<th>Provider</th><th>Max Consecutive</th><th>Visual</th>'
             '<th>WK Assigned</th><th>WE Assigned</th><th>Sites</th></tr></thead><tbody>')

    stretch_sorted = sorted(provider_summary.items(), key=lambda x: -x[1]["max_consecutive_days"])
    for pname, ps in stretch_sorted:
        mc = ps["max_consecutive_days"]
        if mc == 0:
            continue

        # Color-coded bar
        if mc > 12:
            bar_color = "#b71c1c"
        elif mc >= 8:
            bar_color = "#e65100"
        else:
            bar_color = "#2e7d32"
        bar_width = min(mc * 7, 100)

        site_list = ", ".join(SITE_SHORT.get(s, s[:3]) for s in sorted(ps["site_distribution"].keys()))

        h.append(f'<tr><td><a class="prov-link" onclick="showProvider(\'{prov_id(pname)}\')">{esc(pname)}</a></td>')
        h.append(f'<td>{mc}</td>')
        h.append(f'<td><div class="stretch-bar" style="width:{bar_width}px;background:{bar_color}"></div> {mc}d</td>')
        h.append(f'<td>{ps["weeks_assigned"]}</td><td>{ps["weekends_assigned"]}</td>')
        h.append(f'<td>{site_list}</td></tr>')

    h.append('</tbody></table>')

    # Stretch pairing analysis
    h.append('<h2>Stretch Pairing (Same-Site Week+Weekend)</h2>')
    pair_same = 0
    pair_diff = 0
    pair_none = 0
    for period in draft_schedule:
        pidx = period["period_idx"]
        p = periods[pidx]
        if p["type"] != "week":
            continue
        wnum = p["num"]
        # Find matching weekend
        we_period = None
        for pd2 in draft_schedule:
            p2 = periods[pd2["period_idx"]]
            if p2["type"] == "weekend" and p2["num"] == wnum:
                we_period = pd2
                break
        if not we_period:
            continue

        wk_provs = {a["provider"]: a["site"] for a in period["assignments"]}
        we_provs = {a["provider"]: a["site"] for a in we_period["assignments"]}

        for prov in we_provs:
            if prov in wk_provs:
                if wk_provs[prov] == we_provs[prov]:
                    pair_same += 1
                else:
                    pair_diff += 1
            else:
                pair_none += 1

    total_pairs = pair_same + pair_diff + pair_none
    same_pct = _pct(pair_same, total_pairs) if total_pairs > 0 else 0

    h.append(f"""<div class="stats-grid">
<div class="stat-card green"><div class="stat-value">{same_pct:.1f}%</div><div class="stat-label">Same-Site Pairs</div></div>
<div class="stat-card"><div class="stat-value">{pair_same}</div><div class="stat-label">Same Site</div></div>
<div class="stat-card"><div class="stat-value">{pair_diff}</div><div class="stat-label">Different Site</div></div>
<div class="stat-card"><div class="stat-value">{pair_none}</div><div class="stat-label">Weekend Only</div></div>
</div>""")

    h.append('</div>')  # end tab5

    # ═══ TAB 6: GAPS ═════════════════════════════════════════════════════
    h.append('<div class="tab-content" id="tab6">')
    h.append('<h2>Gap Report</h2>')

    if not gap_report:
        h.append(f"""<div class="stats-grid">
<div class="stat-card green"><div class="stat-value">0</div><div class="stat-label">Site Gaps</div></div>
<div class="stat-card green"><div class="stat-value">100%</div><div class="stat-label">All Sites Filled</div></div>
</div>
<p style="color:var(--ok-text);font-weight:600;font-size:16px;margin:20px 0">
All sites fully staffed — no gaps to report.</p>""")
    else:
        # Gap summary
        gaps_with_cands = sum(1 for g in gap_report if g.get("candidates"))
        gaps_without = sum(1 for g in gap_report if not g.get("candidates"))

        h.append(f"""<div class="stats-grid">
<div class="stat-card red"><div class="stat-value">{len(gap_report)}</div><div class="stat-label">Total Gaps</div></div>
<div class="stat-card orange"><div class="stat-value">{gaps_with_cands}</div><div class="stat-label">With Candidates</div></div>
<div class="stat-card red"><div class="stat-value">{gaps_without}</div><div class="stat-label">No Candidates</div></div>
</div>""")

        h.append('<table><thead><tr><th>Site</th><th>Week</th><th>Type</th>'
                 '<th>Dates</th><th>Candidates</th></tr></thead><tbody>')

        for gap in gap_report:
            site = gap["site"]
            color = SITE_COLORS.get(site, "#666")
            pidx = gap.get("period_idx", 0)
            p = periods[pidx] if pidx < len(periods) else {}
            dates_str = f"{p['dates'][0]} to {p['dates'][-1]}" if p.get("dates") else "—"
            cands = gap.get("candidates", [])

            h.append(f'<tr><td><span class="site-badge" style="background:{color}">{esc(site)}</span></td>')
            h.append(f'<td>{p.get("num", "?")}</td><td>{p.get("type", "?")}</td><td>{dates_str}</td>')

            if cands:
                cand_html = ", ".join(
                    f'<a class="prov-link" style="display:inline" onclick="showProvider(\'{prov_id(c["provider"])}\')">'
                    f'{esc(c["provider"])}</a> ({c["score"]:.0f})'
                    for c in cands[:5]
                )
                h.append(f'<td>{cand_html}</td>')
            else:
                h.append('<td><span class="badge badge-short">No candidates</span></td>')
            h.append('</tr>')

        h.append('</tbody></table>')

    # Provider utilization gaps — who still owes after B3
    h.append('<h2>Providers Still Owing After Block 3</h2>')
    h.append('<p style="color:#666;margin-bottom:8px">Providers whose B3 assignment is less than their remaining '
             'annual obligation. They will need to make up these weeks/weekends.</p>')

    util_gaps = [(p, ps) for p, ps in provider_summary.items()
                 if ps["weeks_gap"] + ps["weekends_gap"] >= 1]

    if util_gaps:
        h.append(f'<p style="margin-bottom:8px"><strong>{len(util_gaps)}</strong> providers with remaining obligation after B3</p>')
        h.append('<table><thead><tr><th>Provider</th><th>Annual WK</th><th>B1+B2</th>'
                 '<th>B3 Owed</th><th>B3 Asgn</th><th>Still Owed</th>'
                 '<th>Annual WE</th><th>B1+B2</th><th>B3 Owed</th>'
                 '<th>B3 Asgn</th><th>Still Owed</th><th>Total Gap</th></tr></thead><tbody>')

        for p, ps in sorted(util_gaps, key=lambda x: -(x[1]["weeks_gap"] + x[1]["weekends_gap"])):
            total = ps["weeks_gap"] + ps["weekends_gap"]
            row_style = ' style="background:#fce4ec"' if total >= 5 else ""
            wk_prior = ps.get("weeks_prior", 0)
            we_prior = ps.get("weekends_prior", 0)

            h.append(f'<tr{row_style}><td>{esc(p)}</td>')
            h.append(f'<td>{ps["annual_weeks"]}</td><td>{wk_prior}</td>')
            h.append(f'<td>{ps["weeks_target"]}</td><td>{ps["weeks_assigned"]}</td><td>{ps["weeks_gap"]}</td>')
            h.append(f'<td>{ps["annual_weekends"]}</td><td>{we_prior}</td>')
            h.append(f'<td>{ps["weekends_target"]}</td><td>{ps["weekends_assigned"]}</td><td>{ps["weekends_gap"]}</td>')
            h.append(f'<td><span class="badge badge-short">{total}</span></td></tr>')

        h.append('</tbody></table>')
    else:
        h.append('<p style="color:var(--ok-text);font-weight:600">All providers fully scheduled for remaining annual obligation.</p>')

    h.append('</div>')  # end tab6

    # ═══ JAVASCRIPT ══════════════════════════════════════════════════════
    h.append("""
<script>
function switchTab(idx) {
  document.querySelectorAll('.tab').forEach(function(t, i) {
    t.classList.toggle('active', i === idx);
  });
  document.querySelectorAll('.tab-content').forEach(function(tc, i) {
    tc.classList.toggle('active', i === idx);
  });
}

function toggleCard(header) {
  var card = header.parentElement;
  card.classList.toggle('open');
}

function showProvider(pid) {
  switchTab(3);
  var card = document.getElementById(pid);
  if (card) {
    card.classList.add('open');
    setTimeout(function() { card.scrollIntoView({behavior: 'smooth', block: 'center'}); }, 100);
  }
}

function filterProviders() {
  var q = document.getElementById('provSearch').value.toUpperCase();
  document.querySelectorAll('.prov-card').forEach(function(card) {
    card.style.display = card.textContent.toUpperCase().includes(q) ? '' : 'none';
  });
}

function filterSite(site, btn) {
  document.querySelectorAll('.filter-btn').forEach(function(b) { b.classList.remove('active'); });
  btn.classList.add('active');
  document.querySelectorAll('.site-col').forEach(function(col) {
    if (site === 'all') {
      col.style.display = '';
    } else {
      col.style.display = col.getAttribute('data-site') === site ? '' : 'none';
    }
  });
}

function sortUtil(colIdx, type, th) {
  var table = document.getElementById('utilTable');
  var tbody = table.querySelector('tbody');
  var rows = Array.from(tbody.querySelectorAll('tr'));
  var asc = th.dataset.sortDir !== 'asc';

  // Clear other sort indicators
  table.querySelectorAll('th').forEach(function(t) {
    t.classList.remove('sort-asc', 'sort-desc');
    delete t.dataset.sortDir;
  });

  th.dataset.sortDir = asc ? 'asc' : 'desc';
  th.classList.add(asc ? 'sort-asc' : 'sort-desc');

  rows.sort(function(a, b) {
    var va = a.cells[colIdx].textContent.trim();
    var vb = b.cells[colIdx].textContent.trim();
    if (type === 'num') {
      va = parseFloat(va.replace('%','')) || 0;
      vb = parseFloat(vb.replace('%','')) || 0;
    }
    if (va < vb) return asc ? -1 : 1;
    if (va > vb) return asc ? 1 : -1;
    return 0;
  });

  rows.forEach(function(r) { tbody.appendChild(r); });
}
</script>
""")

    h.append("</div>")  # end container
    h.append("</div>")  # end report-content

    # Password unlock script
    if pw_hash:
        h.append(f"""<script>
(function() {{
  var HASH = "{pw_hash}";
  var overlay = document.getElementById("pw-overlay");
  var content = document.getElementById("report-content");
  var input = document.getElementById("pw-input");
  var btn = document.getElementById("pw-btn");
  var errEl = document.getElementById("pw-error");
  if (!overlay) return;

  async function sha256(str) {{
    var buf = new TextEncoder().encode(str);
    var hash = await crypto.subtle.digest("SHA-256", buf);
    return Array.from(new Uint8Array(hash)).map(function(b) {{
      return b.toString(16).padStart(2, "0");
    }}).join("");
  }}

  async function tryUnlock() {{
    var pw = input.value;
    var h = await sha256(pw);
    if (h === HASH) {{
      overlay.classList.add("hidden");
      content.classList.add("unlocked");
    }} else {{
      errEl.style.display = "block";
      input.value = "";
      input.focus();
    }}
  }}

  btn.addEventListener("click", tryUnlock);
  input.addEventListener("keydown", function(e) {{
    if (e.key === "Enter") tryUnlock();
  }});
  input.focus();
}})();
</script>""")

    h.append("</body></html>")

    os.makedirs(output_dir, exist_ok=True)
    path = os.path.join(output_dir, filename)
    with open(path, "w") as f:
        f.write("\n".join(h))
    print(f"  Saved HTML report: {path}")
    return path


def _build_nav_html(all_results, active_seed=None, active_page=None):
    """Build navigation HTML linking all seed variations, index, and inputs.

    Args:
        all_results: list of engine result dicts
        active_seed: seed number for the active per-seed page (None if not a seed page)
        active_page: 'overview', 'inputs', or None (for seed pages)
    """
    seeds = [r["stats"]["seed"] for r in all_results]
    parts = ['<div class="nav"><span class="nav-label">Variations:</span>']
    # Overview link
    idx_cls = ' class="active"' if active_page == "overview" else ''
    parts.append(f'<a href="v3_index.html"{idx_cls}>Overview</a>')
    # Per-seed links
    for s in seeds:
        cls = ' class="active"' if s == active_seed else ''
        parts.append(f'<a href="report_seed{s}.html"{cls}>Seed {s}</a>')
    # Inputs link
    inp_cls = ' class="active"' if active_page == "inputs" else ''
    parts.append(f'<a href="v3_inputs.html"{inp_cls}>Inputs</a>')
    parts.append('</div>')
    return "".join(parts)


def generate_multi_seed_report(all_results, output_dir):
    """Generate reports for multiple seeds plus a comparison index.

    Args:
        all_results: list of engine output dicts (one per seed)
        output_dir: directory to write HTML files
    """
    os.makedirs(output_dir, exist_ok=True)

    # Generate per-seed reports with cross-navigation
    for results in all_results:
        seed = results["stats"]["seed"]
        nav = _build_nav_html(all_results, active_seed=seed)
        generate_report(results, output_dir, filename=f"report_seed{seed}.html", nav_html=nav)

    # Generate index page if multiple seeds
    if len(all_results) > 1:
        _generate_index(all_results, output_dir)

    # Generate inputs reference page
    _generate_inputs_page(all_results, output_dir)


def _generate_index(all_results, output_dir):
    """Generate index page comparing multiple seed variations."""
    r0 = all_results[0]
    periods = r0["periods"]
    all_dates = []
    for p in periods:
        all_dates.extend(p["dates"])
    block_start = datetime.strptime(min(all_dates), "%Y-%m-%d")
    block_end = datetime.strptime(max(all_dates), "%Y-%m-%d")
    date_range = f"{block_start.strftime('%B %d')} &ndash; {block_end.strftime('%B %d, %Y')}"
    now = datetime.now().strftime("%Y-%m-%d %H:%M")

    # Password gate
    pw_hash = _load_password()
    pw_gate_html = ""
    if pw_hash:
        pw_gate_html = ('<div id="pw-overlay"><div id="pw-box">'
                        '<h2>Password Required</h2>'
                        '<p>This report contains protected information.</p>'
                        '<input type="password" id="pw-input" placeholder="Enter password" autocomplete="off">'
                        '<button id="pw-btn">Unlock</button>'
                        '<div id="pw-error">Incorrect password</div>'
                        '</div></div>')

    content_cls = "unlocked" if not pw_hash else ""
    nav_html = _build_nav_html(all_results, active_page="overview")

    h = [f"""<!DOCTYPE html>
<html lang="en"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Block 3 Schedule — V3 Overview</title>
<style>{_common_css()}</style></head><body>

{pw_gate_html}

<div id="report-content" class="{content_cls}">
<div class="container">

{nav_html}

<h1>Block 3 Schedule Report — V3 Engine</h1>
<p class="subtitle">{date_range}</p>
<p class="generated">Generated: {now}</p>

<h2>Seed Comparison</h2>
<p style="color:#666;margin-bottom:12px">Each variation uses a different random seed, producing a different assignment pattern while following the same rules. Click a seed to see its full report.</p>
<table><thead><tr>
<th>Seed</th><th>WK Coverage</th><th>WE Coverage</th><th>Gaps</th>
<th>Zero-Gap Viol</th><th>Weeks Asgn</th><th>WE Asgn</th><th>At Capacity</th><th>Under-Used</th><th>Report</th>
</tr></thead><tbody>
"""]

    for results in all_results:
        s = results["stats"]
        seed = s["seed"]
        wk_cls = "badge-ok" if s["weekday_coverage_pct"] == 100 else "badge-short"
        we_cls = "badge-ok" if s["weekend_coverage_pct"] == 100 else "badge-short"
        gap_cls = "badge-ok" if s["total_gaps"] == 0 else "badge-short"

        h.append(f'<tr><td><strong>{seed}</strong></td>')
        h.append(f'<td><span class="badge {wk_cls}">{s["weekday_coverage_pct"]:.1f}%</span></td>')
        h.append(f'<td><span class="badge {we_cls}">{s["weekend_coverage_pct"]:.1f}%</span></td>')
        h.append(f'<td><span class="badge {gap_cls}">{s["total_gaps"]}</span></td>')
        h.append(f'<td>{s["zero_gap_violations"]}</td>')
        h.append(f'<td>{s["total_weeks_assigned"]}</td><td>{s["total_weekends_assigned"]}</td>')
        h.append(f'<td>{s["providers_at_capacity"]}</td>')
        h.append(f'<td>{s["providers_under_utilized"]}</td>')
        h.append(f'<td><a href="report_seed{seed}.html" style="color:var(--link);font-weight:600">View &rarr;</a></td></tr>')

    h.append('</tbody></table>')
    h.append('</div>')  # end container
    h.append('</div>')  # end report-content

    # Password unlock script
    if pw_hash:
        h.append(f"""<script>
(function() {{
  var HASH = "{pw_hash}";
  var overlay = document.getElementById("pw-overlay");
  var content = document.getElementById("report-content");
  var input = document.getElementById("pw-input");
  var btn = document.getElementById("pw-btn");
  var errEl = document.getElementById("pw-error");
  if (!overlay) return;

  async function sha256(str) {{
    var buf = new TextEncoder().encode(str);
    var hash = await crypto.subtle.digest("SHA-256", buf);
    return Array.from(new Uint8Array(hash)).map(function(b) {{
      return b.toString(16).padStart(2, "0");
    }}).join("");
  }}

  async function tryUnlock() {{
    var pw = input.value;
    var h = await sha256(pw);
    if (h === HASH) {{
      overlay.classList.add("hidden");
      content.classList.add("unlocked");
    }} else {{
      errEl.style.display = "block";
      input.value = "";
      input.focus();
    }}
  }}

  btn.addEventListener("click", tryUnlock);
  input.addEventListener("keydown", function(e) {{
    if (e.key === "Enter") tryUnlock();
  }});
  input.focus();
}})();
</script>""")

    h.append("</body></html>")

    path = os.path.join(output_dir, "v3_index.html")
    with open(path, "w") as f:
        f.write("\n".join(h))
    print(f"  Saved index: {path}")


def _generate_inputs_page(all_results, output_dir):
    """Generate an inputs reference page showing provider data from the Excel sheet."""
    # Load Excel data
    v3_dir = os.path.dirname(os.path.abspath(__file__))
    excel_path = os.path.join(v3_dir, "input", "hospitalist_scheduler.xlsx")
    if not os.path.exists(excel_path) or _load_workbook is None:
        print("  Skipping inputs page (Excel not found or openpyxl not available)")
        return

    wb = _load_workbook(excel_path, data_only=True)
    providers = load_providers_from_excel(wb)
    tags_data = load_tags_from_excel(wb)
    sites_demand = load_sites_from_excel(wb)
    wb.close()

    now = datetime.now().strftime("%Y-%m-%d %H:%M")

    # Password gate
    pw_hash = _load_password()
    pw_gate_html = ""
    if pw_hash:
        pw_gate_html = ('<div id="pw-overlay"><div id="pw-box">'
                        '<h2>Password Required</h2>'
                        '<p>This report contains protected information.</p>'
                        '<input type="password" id="pw-input" placeholder="Enter password" autocomplete="off">'
                        '<button id="pw-btn">Unlock</button>'
                        '<div id="pw-error">Incorrect password</div>'
                        '</div></div>')

    content_cls = "unlocked" if not pw_hash else ""
    nav_html = _build_nav_html(all_results, active_page="inputs")

    h = [f"""<!DOCTYPE html>
<html lang="en"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Block 3 Schedule — V3 Inputs</title>
<style>{_common_css()}</style></head><body>

{pw_gate_html}

<div id="report-content" class="{content_cls}">
<div class="container">

{nav_html}

<h1>Block 3 Inputs Reference — V3 Engine</h1>
<p class="generated">Generated: {now}</p>

<div class="tabs">
  <div class="tab active" onclick="switchTab(0)">Providers</div>
  <div class="tab" onclick="switchTab(1)">Site Demand</div>
  <div class="tab" onclick="switchTab(2)">Tags</div>
</div>
"""]

    # ─── Tab 0: Providers ─────────────────────────────────────────
    sorted_provs = sorted(providers.keys())
    total = len(sorted_provs)
    days_count = sum(1 for p in providers.values() if p["shift_type"] == "Days")
    nights_count = sum(1 for p in providers.values() if p["shift_type"] == "Nights")
    hybrid_count = sum(1 for p in providers.values() if p["shift_type"] == "Hybrid")
    total_wk_remaining = sum(p["weeks_remaining"] for p in providers.values())
    total_we_remaining = sum(p["weekends_remaining"] for p in providers.values())

    h.append(f"""<div class="tab-content active">
<div class="stats-grid">
<div class="stat-card"><div class="stat-value">{total}</div><div class="stat-label">Total Providers</div></div>
<div class="stat-card"><div class="stat-value">{days_count}</div><div class="stat-label">Day Shift</div></div>
<div class="stat-card"><div class="stat-value">{nights_count}</div><div class="stat-label">Night Shift</div></div>
<div class="stat-card"><div class="stat-value">{hybrid_count}</div><div class="stat-label">Hybrid</div></div>
<div class="stat-card"><div class="stat-value">{total_wk_remaining:.0f}</div><div class="stat-label">Total WK Remaining</div></div>
<div class="stat-card"><div class="stat-value">{total_we_remaining:.0f}</div><div class="stat-label">Total WE Remaining</div></div>
</div>

<input type="text" class="search-box" placeholder="Search providers..." oninput="filterInputProviders(this.value)">

<table id="inputsTable">
<thead><tr>
<th>Provider</th><th>Shift</th><th>FTE</th>
<th>Ann WK</th><th>Ann WE</th>
<th>B1+B2 WK</th><th>B1+B2 WE</th>
<th>Rem WK</th><th>Rem WE</th>
<th>Cooper%</th><th>VEB%</th><th>MHW%</th>
<th>Mann%</th><th>Virtua%</th><th>Cape%</th>
<th>Eligible Sites</th><th>Tags</th>
</tr></thead><tbody>
""")

    for pname in sorted_provs:
        pd = providers[pname]
        eligible = get_eligible_sites(pname, pd, tags_data)
        ptags = tags_data.get(pname, [])

        # Site badges
        site_badges = " ".join(
            f'<span class="site-badge" style="background:{SITE_COLORS.get(s, "#666")}">'
            f'{SITE_SHORT.get(s, s[:3])}</span>'
            for s in eligible
        )

        # Tags
        tag_strs = []
        for t in ptags:
            tag_strs.append(esc(t["tag"]) + (f': {esc(t["rule"])}' if t["rule"] else ''))
        tag_html = '<br>'.join(tag_strs) if tag_strs else ''

        # Shift type color
        shift_cls = ""
        if pd["shift_type"] == "Nights":
            shift_cls = ' style="background:#e8eaf6"'
        elif pd["shift_type"] == "Hybrid":
            shift_cls = ' style="background:#fff3e0"'

        h.append(f'<tr{shift_cls}>')
        h.append(f'<td><strong>{esc(pname)}</strong></td>')
        h.append(f'<td>{esc(pd["shift_type"])}</td>')
        h.append(f'<td>{pd["fte"]:.2f}</td>')
        h.append(f'<td>{pd["annual_weeks"]:.0f}</td><td>{pd["annual_weekends"]:.0f}</td>')
        h.append(f'<td>{pd["prior_weeks_worked"]:.0f}</td><td>{pd["prior_weekends_worked"]:.0f}</td>')
        h.append(f'<td>{pd["weeks_remaining"]:.0f}</td>')
        h.append(f'<td>{pd["weekends_remaining"]:.0f}</td>')
        h.append(f'<td>{pd["pct_cooper"]:.0f}</td><td>{pd["pct_inspira_veb"]:.0f}</td>')
        h.append(f'<td>{pd["pct_inspira_mhw"]:.0f}</td><td>{pd["pct_mannington"]:.0f}</td>')
        h.append(f'<td>{pd["pct_virtua"]:.0f}</td><td>{pd["pct_cape"]:.0f}</td>')
        h.append(f'<td>{site_badges}</td>')
        h.append(f'<td style="font-size:11px">{tag_html}</td>')
        h.append('</tr>')

    h.append('</tbody></table>')
    h.append('</div>')  # end tab 0

    # ─── Tab 1: Site Demand ─────────────────────────────────────
    h.append('<div class="tab-content">')
    h.append('<h2>Site Staffing Demand</h2>')
    h.append('<p style="color:#666;margin-bottom:12px">Providers needed per site per period type, '
             'as defined in the Excel "Sites" sheet.</p>')

    sites_grouped = {}
    for (site, dtype), needed in sorted(sites_demand.items()):
        if site not in sites_grouped:
            sites_grouped[site] = {}
        sites_grouped[site][dtype] = needed

    h.append('<table><thead><tr><th>Site</th><th>Weekday</th><th>Weekend</th>'
             '<th>Total / Week</th></tr></thead><tbody>')
    total_wd = 0
    total_we_d = 0
    for site in sorted(sites_grouped.keys()):
        d = sites_grouped[site]
        wd = d.get("weekday", 0)
        we = d.get("weekend", 0)
        total_wd += wd
        total_we_d += we
        color = SITE_COLORS.get(site, "#666")
        h.append(f'<tr><td><span class="site-badge" style="background:{color}">{esc(site)}</span></td>')
        h.append(f'<td>{wd}</td><td>{we}</td><td>{wd + we}</td></tr>')

    h.append(f'<tr style="font-weight:700;background:#e8eaf6"><td>Total</td>'
             f'<td>{total_wd}</td><td>{total_we_d}</td><td>{total_wd + total_we_d}</td></tr>')
    h.append('</tbody></table>')

    # Provider pool per site
    h.append('<h2>Provider Pool by Site</h2>')
    h.append('<p style="color:#666;margin-bottom:12px">Eligible providers per site '
             '(based on allocation percentages and tag restrictions).</p>')
    h.append('<table><thead><tr><th>Site</th><th>Eligible Providers</th>'
             '<th>Ratio (Pool / Demand)</th></tr></thead><tbody>')
    for site in sorted(sites_grouped.keys()):
        count = sum(1 for pname in providers
                    if site in get_eligible_sites(pname, providers[pname], tags_data))
        wd = sites_grouped[site].get("weekday", 1)
        ratio = count / wd if wd > 0 else 0
        color = SITE_COLORS.get(site, "#666")
        ratio_cls = "badge-ok" if ratio >= 3 else ("badge-warn" if ratio >= 2 else "badge-short")
        h.append(f'<tr><td><span class="site-badge" style="background:{color}">{esc(site)}</span></td>')
        h.append(f'<td>{count}</td>')
        h.append(f'<td><span class="badge {ratio_cls}">{ratio:.1f}x</span></td></tr>')
    h.append('</tbody></table>')
    h.append('</div>')  # end tab 1

    # ─── Tab 2: Tags ─────────────────────────────────────────────
    h.append('<div class="tab-content">')
    h.append('<h2>Provider Tags</h2>')
    h.append('<p style="color:#666;margin-bottom:12px">Tags modify scheduling behavior '
             'for individual providers.</p>')

    tag_summary = {}
    for pname, ptags in sorted(tags_data.items()):
        for t in ptags:
            tag_name = t["tag"]
            if tag_name not in tag_summary:
                tag_summary[tag_name] = []
            tag_summary[tag_name].append({"provider": pname, "rule": t["rule"]})

    if tag_summary:
        h.append('<div class="stats-grid">')
        h.append(f'<div class="stat-card"><div class="stat-value">{len(tag_summary)}</div>'
                 f'<div class="stat-label">Unique Tags</div></div>')
        h.append(f'<div class="stat-card"><div class="stat-value">'
                 f'{sum(len(v) for v in tag_summary.values())}</div>'
                 f'<div class="stat-label">Total Assignments</div></div>')
        h.append(f'<div class="stat-card"><div class="stat-value">{len(tags_data)}</div>'
                 f'<div class="stat-label">Providers w/ Tags</div></div>')
        h.append('</div>')

        for tag_name in sorted(tag_summary.keys()):
            entries = tag_summary[tag_name]
            h.append(f'<h3>{esc(tag_name)} <span class="badge badge-warn">{len(entries)}</span></h3>')
            h.append('<table><thead><tr><th>Provider</th><th>Rule</th></tr></thead><tbody>')
            for e in sorted(entries, key=lambda x: x["provider"]):
                h.append(f'<tr><td>{esc(e["provider"])}</td><td>{esc(e["rule"]) or "—"}</td></tr>')
            h.append('</tbody></table>')
    else:
        h.append('<p>No tags defined.</p>')

    h.append('</div>')  # end tab 2

    # ─── JavaScript ───────────────────────────────────────────────
    h.append("""
<script>
function switchTab(idx) {
  document.querySelectorAll('.tab').forEach(function(t, i) {
    t.classList.toggle('active', i === idx);
  });
  document.querySelectorAll('.tab-content').forEach(function(tc, i) {
    tc.classList.toggle('active', i === idx);
  });
}
function filterInputProviders(q) {
  q = q.toUpperCase();
  document.querySelectorAll('#inputsTable tbody tr').forEach(function(row) {
    row.style.display = row.textContent.toUpperCase().includes(q) ? '' : 'none';
  });
}
</script>
""")

    h.append('</div>')  # end container
    h.append('</div>')  # end report-content

    # Password unlock script
    if pw_hash:
        h.append(f"""<script>
(function() {{
  var HASH = "{pw_hash}";
  var overlay = document.getElementById("pw-overlay");
  var content = document.getElementById("report-content");
  var input = document.getElementById("pw-input");
  var btn = document.getElementById("pw-btn");
  var errEl = document.getElementById("pw-error");
  if (!overlay) return;

  async function sha256(str) {{
    var buf = new TextEncoder().encode(str);
    var hash = await crypto.subtle.digest("SHA-256", buf);
    return Array.from(new Uint8Array(hash)).map(function(b) {{
      return b.toString(16).padStart(2, "0");
    }}).join("");
  }}

  async function tryUnlock() {{
    var pw = input.value;
    var h = await sha256(pw);
    if (h === HASH) {{
      overlay.classList.add("hidden");
      content.classList.add("unlocked");
    }} else {{
      errEl.style.display = "block";
      input.value = "";
      input.focus();
    }}
  }}

  btn.addEventListener("click", tryUnlock);
  input.addEventListener("keydown", function(e) {{
    if (e.key === "Enter") tryUnlock();
  }});
  input.focus();
}})();
</script>""")

    h.append("</body></html>")

    path = os.path.join(output_dir, "v3_inputs.html")
    with open(path, "w") as f:
        f.write("\n".join(h))
    print(f"  Saved inputs: {path}")
