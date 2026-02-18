#!/usr/bin/env python3
"""
Recalculate prior blocks worked (Blocks 1 & 2) from the HTML schedule files.

Reads monthly HTML schedules from input/ for Jun 2025 through Feb 2026,
classifies services per docs/block-scheduling-rules.md Section 1.3,
filters to Block 1 & 2 date boundaries, and calculates:
  prior_weeks    = weekday_shifts / 5   (decimal, rounded to 1)
  prior_weekends = weekend_shifts / 2   (decimal, rounded to 1)
  prior_nights   = night_shifts         (raw count)

Outputs:
  output/prior_actuals.json  — full results for engine consumption
  output/prior_actuals.xlsx  — single sheet for Google Sheet paste
"""

import csv
import io
import json
import math
import os
import re
import sys
import urllib.request
from collections import defaultdict
from datetime import datetime, date

# Add project root to path
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(SCRIPT_DIR)
sys.path.insert(0, PROJECT_ROOT)

from parse_schedule import parse_schedule, merge_schedules
from name_match import to_canonical, normalize_name, match_provider, clean_html_provider

INPUT_DIR = os.path.join(PROJECT_ROOT, "input")
OUTPUT_DIR = os.path.join(PROJECT_ROOT, "output")

# Block boundaries (inclusive)
BLOCK_1_START = date(2025, 6, 30)
BLOCK_1_END = date(2025, 11, 2)
BLOCK_2_START = date(2025, 11, 3)
BLOCK_2_END = date(2026, 3, 1)

# Files covering Blocks 1 & 2 (June 2025 through Feb 2026)
PRIOR_FILES = [
    "Hospital Medicine Schedule, 6_1 to 6_30, 2025.html",
    "Hospital Medicine Schedule, 7_1 to 7_31, 2025.html",
    "Hospital Medicine Schedule, 8_1 to 8_31, 2025.html",
    "Hospital Medicine Schedule, 9_1 to 9_30, 2025.html",
    "Hospital Medicine Schedule, 10_1 to 10_31, 2025.html",
    "Hospital Medicine Schedule, 11_1 to 11_30, 2025.html",
    "Hospital Medicine Schedule, 12_1 to 12_31, 2025.html",
    "Hospital Medicine Schedule, 1_1 to 1_31, 2026.html",
    "Hospital Medicine Schedule, 2_1 to 2_28, 2026.html",
]


# ---------------------------------------------------------------------------
# Service Classification
# ---------------------------------------------------------------------------
# Rules from docs/block-scheduling-rules.md Section 1.3
# Returns: "exclude", "day", "night", or "swing"

def classify_service(service_name, hours):
    """Classify a service as 'day', 'night', 'swing', or 'exclude'."""
    sname = service_name.lower()
    hours_lower = hours.lower() if hours else ""

    # ── INCLUDE overrides (these match exclusion patterns but are included) ──
    # CCPA / Physician Advisor — physician day shift
    if "ccpa" in sname or "physician advisor" in sname:
        return "day"
    # Hospital Medicine Consults (physician, not APP)
    if sname == "hospital medicine consults":
        return "day"
    # Inspira Mannington PA — physician shift despite "PA" in name
    if "mannington pa" in sname and "app" not in sname:
        return "day"
    # Cape PA — physician shift, not APP (per rules doc Section 1.3)
    if sname == "cape pa":
        return "day"
    # UM Referrals / UM Rounds — physician shifts
    if "um referral" in sname or "um rounds" in sname:
        return "day"

    # ── EXCLUDE rules (priority order from rules doc) ─────────────────

    # APP roles
    if " app" in sname or sname.startswith("app ") or sname.endswith(" app"):
        return "exclude"
    if "apn" in sname:
        return "exclude"
    if " pa " in sname or sname.endswith(" pa"):
        return "exclude"

    # Night Coverage (APP/resident coverage, not physician nocturnist)
    if "night coverage" in sname:
        return "exclude"

    # Resident / Fellow
    if "resident" in sname:
        return "exclude"
    if "fellow" in sname:
        return "exclude"

    # Behavioral medicine
    if "behavioral" in sname:
        return "exclude"

    # Site Director
    if "site director" in sname:
        return "exclude"

    # Admin
    if "admin" in sname:
        return "exclude"

    # Hospice
    if "hospice" in sname:
        return "exclude"

    # Kessler Rehab
    if "kessler" in sname:
        return "exclude"

    # Holy Redeemer
    if "holy redeemer" in sname:
        return "exclude"

    # Cape RMD
    if "cape rmd" in sname:
        return "exclude"

    # Long Call (all variants)
    if "long call" in sname:
        return "exclude"
    if sname.startswith("direct care long call"):
        return "exclude"

    # Virtua Coverage (only when BOTH "virtua" AND "coverage" present)
    if "virtua" in sname and "coverage" in sname:
        return "exclude"

    # UM — physician shift, counts as day work (per rules doc Section 1.3)
    # UM Referrals / UM Rounds are also included (handled in overrides above)
    if sname.strip() == "um":
        return "day"

    # Consults (remaining — Hospital Medicine Consults handled above)
    if "consult" in sname:
        return "exclude"

    # Moonlighting service names (resident moonlighters)
    if "moonlighting" in sname:
        return "exclude"

    # LTC-SAR APP variants (but NOT "Cape LTC-SAR on call Physician")
    if "ltc-sar" in sname and "physician" not in sname:
        return "exclude"

    # DO NOT USE
    if "do not use" in sname:
        return "exclude"

    # CC (standalone admin-like)
    if sname.strip() == "cc":
        return "exclude"

    # APP Lead
    if "app lead" in sname:
        return "exclude"

    # Teaching Admissions (short shifts, APP-like)
    if "teaching admissions" in sname:
        return "exclude"

    # Night APP
    if "night app" in sname:
        return "exclude"

    # ── NIGHT shifts ──────────────────────────────────────────────────
    if any(kw in sname for kw in ["night", "nocturnist", "(nah)"]):
        return "night"

    # Hours-based night detection
    if re.match(r'[57]p-[57]a', hours_lower):
        return "night"
    if re.match(r'5a-7a', hours_lower):
        return "night"
    if re.match(r'11p-7a', hours_lower):
        return "night"

    # ── SWING shifts ──────────────────────────────────────────────────
    if "swing" in sname:
        return "swing"
    if re.match(r'[1-4]p-', hours_lower):
        return "swing"

    # ── DAY shifts — everything remaining ─────────────────────────────
    return "day"


def parse_date(date_str):
    """Parse M/D/YYYY date string from parsed HTML schedule."""
    try:
        parts = date_str.split("/")
        if len(parts) == 3:
            return date(int(parts[2]), int(parts[0]), int(parts[1]))
    except (ValueError, IndexError):
        pass
    return None


def in_block_range(d):
    """Check if a date falls within Block 1 or Block 2."""
    if d is None:
        return False
    return (BLOCK_1_START <= d <= BLOCK_1_END) or (BLOCK_2_START <= d <= BLOCK_2_END)


def main():
    print("=" * 70)
    print("RECALCULATING PRIOR BLOCK ACTUALS FROM HTML SCHEDULES")
    print(f"Block 1: {BLOCK_1_START} to {BLOCK_1_END}")
    print(f"Block 2: {BLOCK_2_START} to {BLOCK_2_END}")
    print("=" * 70)

    # Parse all prior month HTML files
    all_month_data = []
    for fname in PRIOR_FILES:
        fpath = os.path.join(INPUT_DIR, fname)
        if not os.path.exists(fpath):
            print(f"  WARNING: Missing file: {fname}")
            continue
        print(f"  Parsing: {fname}")
        month_data = parse_schedule(fpath)
        all_month_data.append(month_data)

    if not all_month_data:
        print("ERROR: No schedule files found!")
        sys.exit(1)

    # Merge all months
    merged = merge_schedules(all_month_data)
    print(f"\nParsed {len(merged['schedule'])} days, "
          f"{len(merged.get('by_provider', {}))} unique providers")

    # Classify all services
    service_classes = {}
    for s in merged["services"]:
        cls = classify_service(s["name"], s["hours"])
        service_classes[s["name"]] = cls

    # Report classification counts
    counts_by_class = defaultdict(int)
    for cls in service_classes.values():
        counts_by_class[cls] += 1
    print(f"\nService classification: {counts_by_class['day']} day, "
          f"{counts_by_class['night']} night, {counts_by_class['swing']} swing, "
          f"{counts_by_class['exclude']} excluded")

    # Check for any services not in the documented lists (flag for review)
    unknown_services = []
    for svc_name, cls in sorted(service_classes.items()):
        if cls != "exclude":
            # This is an included service — should be in the included list
            pass

    # Count shifts per provider, DEDUPLICATED per day
    # Priority: night(3) > swing(2) > weekday/weekend(1)
    provider_day_shifts = defaultdict(dict)  # provider -> {date_str: shift_type}
    dates_outside_range = 0
    moonlighting_excluded = 0

    for day_entry in merged["schedule"]:
        dow = day_entry["day_of_week"]
        is_weekend = dow in ("Sat", "Sun")
        date_str = day_entry["date"]

        # Date filtering: only count shifts within Block 1 & 2 boundaries
        d = parse_date(date_str)
        if not in_block_range(d):
            dates_outside_range += 1
            continue

        for assignment in day_entry["assignments"]:
            # Clean provider name
            provider = clean_html_provider(assignment["provider"])
            if not provider:
                continue

            # Exclude moonlighting shifts
            if assignment.get("moonlighting", False):
                moonlighting_excluded += 1
                continue

            service_name = assignment["service"]
            svc_class = service_classes.get(service_name, "day")

            if svc_class == "exclude":
                continue

            # Determine shift type for this day
            if svc_class == "night":
                shift_type = "night"
            elif svc_class == "swing":
                shift_type = "swing" if is_weekend else "swing"
            elif is_weekend:
                shift_type = "weekend"
            else:
                shift_type = "weekday"

            # Keep highest-priority shift type for this provider+date
            existing = provider_day_shifts[provider].get(date_str)
            priority = {"night": 3, "swing": 2, "weekday": 1, "weekend": 1}
            if existing is None or priority.get(shift_type, 0) > priority.get(existing, 0):
                provider_day_shifts[provider][date_str] = shift_type

    print(f"\nDate filtering: {dates_outside_range} days outside Block 1 & 2 range skipped")
    print(f"Moonlighting shifts excluded: {moonlighting_excluded}")

    # Aggregate from deduplicated day-level data
    provider_counts = defaultdict(lambda: {
        "weekday_shifts": 0,
        "weekend_shifts": 0,
        "night_shifts": 0,
        "swing_shifts": 0,
    })

    for provider, day_map in provider_day_shifts.items():
        for date_str, shift_type in day_map.items():
            if shift_type == "night":
                provider_counts[provider]["night_shifts"] += 1
            elif shift_type == "swing":
                provider_counts[provider]["swing_shifts"] += 1
            elif shift_type == "weekend":
                provider_counts[provider]["weekend_shifts"] += 1
            else:
                provider_counts[provider]["weekday_shifts"] += 1

    # Calculate weeks/weekends/nights using CANONICAL names
    results = {}
    for provider in sorted(provider_counts.keys()):
        counts = provider_counts[provider]
        canonical = to_canonical(provider)
        # If multiple HTML names map to the same canonical, sum them
        if canonical in results:
            results[canonical]["weekday_shifts"] += counts["weekday_shifts"]
            results[canonical]["weekend_shifts"] += counts["weekend_shifts"]
            results[canonical]["night_shifts"] += counts["night_shifts"]
            results[canonical]["swing_shifts"] += counts["swing_shifts"]
            results[canonical]["prior_weeks"] = round(
                results[canonical]["weekday_shifts"] / 5, 1)
            results[canonical]["prior_weekends"] = round(
                results[canonical]["weekend_shifts"] / 2, 1)
            results[canonical]["prior_nights"] = results[canonical]["night_shifts"]
        else:
            results[canonical] = {
                "weekday_shifts": counts["weekday_shifts"],
                "prior_weeks": round(counts["weekday_shifts"] / 5, 1),
                "weekend_shifts": counts["weekend_shifts"],
                "prior_weekends": round(counts["weekend_shifts"] / 2, 1),
                "night_shifts": counts["night_shifts"],
                "prior_nights": counts["night_shifts"],
                "swing_shifts": counts["swing_shifts"],
            }

    # Write JSON output
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    json_path = os.path.join(OUTPUT_DIR, "prior_actuals.json")
    with open(json_path, "w") as f:
        json.dump(results, f, indent=2, sort_keys=True)
    print(f"\nWrote {len(results)} providers to: {json_path}")

    # Write XLSX output
    write_xlsx(results)

    # Print summary table
    print(f"\n{'='*90}")
    print(f"{'Provider':<45s} {'WkDay':>6s} {'Wks':>6s} {'WkEnd':>6s} {'WEs':>6s} "
          f"{'Nights':>6s} {'Swing':>6s}")
    print("-" * 90)
    active_count = 0
    for provider in sorted(results.keys()):
        r = results[provider]
        total = (r["weekday_shifts"] + r["weekend_shifts"]
                 + r["night_shifts"] + r["swing_shifts"])
        if total == 0:
            continue
        active_count += 1
        print(f"{provider:<45s} {r['weekday_shifts']:>6d} {r['prior_weeks']:>6.1f} "
              f"{r['weekend_shifts']:>6d} {r['prior_weekends']:>6.1f} "
              f"{r['night_shifts']:>6d} {r['swing_shifts']:>6d}")

    print(f"\n{active_count} providers with shifts (of {len(results)} total)")

    return results


SHEET_ID = "1dbHUkE-pLtQJK02ig2EbU2eny33N60GQUvk4muiXW5M"
SHEET_BASE = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/gviz/tq?tqx=out:csv"


def _fetch_sheet_provider_order():
    """Fetch provider names from Google Sheet Providers tab to match its row order.
    Falls back to local CSV if the fetch fails."""
    # Try Google Sheet first
    try:
        url = f"{SHEET_BASE}&sheet={urllib.request.quote('Providers')}"
        print("  Fetching provider list from Google Sheet...")
        with urllib.request.urlopen(url, timeout=10) as resp:
            text = resp.read().decode("utf-8")
        reader = csv.DictReader(io.StringIO(text))
        names = []
        for row in reader:
            name = row.get("provider_name", "").strip()
            if name:
                names.append(name)
        if names:
            # Cache locally for offline use
            cache_path = os.path.join(OUTPUT_DIR, "providers_sheet.csv")
            with open(cache_path, "w", newline="") as f:
                f.write(text)
            print(f"  Fetched {len(names)} providers from Google Sheet (cached to providers_sheet.csv)")
            return names
    except Exception as e:
        print(f"  WARNING: Could not fetch Google Sheet: {e}")

    # Fall back to local CSV
    for fname in ("providers_sheet_updated.csv", "providers_sheet.csv"):
        csv_path = os.path.join(OUTPUT_DIR, fname)
        if os.path.isfile(csv_path):
            with open(csv_path, newline="") as f:
                reader = csv.DictReader(f)
                names = []
                for row in reader:
                    name = row.get("provider_name", "").strip()
                    if name:
                        names.append(name)
            if names:
                print(f"  Using local fallback: {fname} ({len(names)} providers)")
                return names
    return None


def write_xlsx(results):
    """Write prior actuals to XLSX, matching Google Sheet provider order.
    Only includes providers present in the Google Sheet."""
    try:
        import openpyxl
    except ImportError:
        print("\nWARNING: openpyxl not installed — skipping XLSX output")
        print("  Install with: pip install openpyxl")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Prior Actuals"

    # Header row
    headers = [
        "provider_name",
        "prior_weeks_worked",
        "prior_weekends_worked",
        "prior_nights_worked",
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(1, col, header)

    # Use Google Sheet order if available, otherwise fall back to alphabetical
    sheet_names = _fetch_sheet_provider_order()
    if sheet_names:
        provider_order = sheet_names
        print(f"  XLSX: matching Google Sheet order ({len(sheet_names)} providers)")
    else:
        provider_order = sorted(results.keys())
        print("  XLSX: Google Sheet CSV not found, using alphabetical order")

    # ═══════════════════════════════════════════════════════════════════════
    # TEMPORARY OVERRIDE — FORCE FAIR-SHARE REMAINING FOR TEST RUN
    # ═══════════════════════════════════════════════════════════════════════
    # These 4 providers have lopsided prior actuals (data quality issues in
    # Blocks 1 & 2) that would skew Block 3 scheduling.  We override their
    # prior_weeks_worked / prior_weekends_worked so the Google Sheet computes
    # remaining = ann − prior = ceil(ann/3)  (i.e. exactly fair-share).
    #
    # Formula:  prior_worked_override = annual − ceil(annual / 3)
    #
    # TO UNDO: Delete this _FAIR_SHARE_OVERRIDES dict and the apply block
    #          below (search for "FAIR-SHARE OVERRIDE" in this file).
    # ═══════════════════════════════════════════════════════════════════════
    _FAIR_SHARE_OVERRIDES = {
        # provider_name (canonical) → (annual_weeks, annual_weekends)
        "GUMMADI, VEDAM":    (24, 20),   # fair-share: 8 wk, 7 we → prior: 16, 13
        "SHKLAR, DAVID":     (17, 12),   # fair-share: 6 wk, 4 we → prior: 11, 8
        "PATTANAIK, SAMBIT": (12, 12),   # fair-share: 4 wk, 4 we → prior: 8, 8
        "PATEL, KAJAL":      (6,  4),    # fair-share: 2 wk, 2 we → prior: 4, 2
    }

    def _fair_share_prior(annual):
        """Compute prior_worked needed so remaining = ceil(annual/3)."""
        return annual - math.ceil(annual / 3)

    # Pre-compute override lookup:  normalized_name → (prior_weeks, prior_weekends)
    _override_lookup = {}
    for oname, (ann_wk, ann_we) in _FAIR_SHARE_OVERRIDES.items():
        _override_lookup[normalize_name(oname)] = (
            _fair_share_prior(ann_wk),
            _fair_share_prior(ann_we),
        )
    # ═══════════════════════════════════════════════════════════════════════

    row = 2
    matched = 0
    missing = 0
    overridden = 0
    for sheet_name in provider_order:
        # ── Check for FAIR-SHARE OVERRIDE first ──────────────────────────
        norm_sheet = normalize_name(sheet_name)
        if norm_sheet in _override_lookup:
            ov_wk, ov_we = _override_lookup[norm_sheet]
            # Look up actual nights from results (not overridden)
            matched_key = match_provider(sheet_name, results.keys())
            actual_nights = results[matched_key]["prior_nights"] if matched_key else 0
            ws.cell(row, 1, sheet_name)
            ws.cell(row, 2, ov_wk)
            ws.cell(row, 3, ov_we)
            ws.cell(row, 4, actual_nights)
            overridden += 1
            matched += 1
            print(f"  ⚠ FAIR-SHARE OVERRIDE: {sheet_name} → "
                  f"prior_weeks={ov_wk}, prior_weekends={ov_we}")
            row += 1
            continue

        # ── Normal path ──────────────────────────────────────────────────
        # Use match_provider for full alias + abbreviation resolution
        matched_key = match_provider(sheet_name, results.keys())
        if matched_key:
            r = results[matched_key]
            ws.cell(row, 1, sheet_name)
            ws.cell(row, 2, r["prior_weeks"])
            ws.cell(row, 3, r["prior_weekends"])
            ws.cell(row, 4, r["prior_nights"])
            matched += 1
        else:
            # Provider in sheet but not in schedule data — output zeros
            ws.cell(row, 1, sheet_name)
            ws.cell(row, 2, 0)
            ws.cell(row, 3, 0)
            ws.cell(row, 4, 0)
            missing += 1
        row += 1

    xlsx_path = os.path.join(OUTPUT_DIR, "prior_actuals.xlsx")
    wb.save(xlsx_path)
    print(f"Wrote XLSX: {xlsx_path} ({matched} matched, {missing} not found in schedules, "
          f"{overridden} fair-share overrides)")


if __name__ == "__main__":
    main()
