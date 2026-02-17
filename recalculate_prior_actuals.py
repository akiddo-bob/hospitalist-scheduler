#!/usr/bin/env python3
"""
Recalculate prior blocks worked (Blocks 1 & 2) from the HTML schedule files.

Reads monthly HTML schedules from input/ for Jul 2025 through Feb 2026,
counts weekday shifts (M-F), weekend shifts (Sa-Su), and night shifts per provider,
then calculates:
  prior_weeks    = weekday_shifts / 5   (decimal, not floored)
  prior_weekends = weekend_shifts / 2   (decimal, not floored)
  prior_nights   = night_shifts / 5     (decimal, not floored)

Updates block_schedule_input.xlsx with the corrected values.
Also writes prior_actuals.json for reference.
"""

import json
import os
import re
import sys
from collections import defaultdict

# Add project root to path
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, SCRIPT_DIR)
from parse_schedule import parse_schedule, merge_schedules

INPUT_DIR = os.path.join(SCRIPT_DIR, "input")
OUTPUT_DIR = os.path.join(SCRIPT_DIR, "output")

# Prior blocks = Jul 2025 through Feb 2026 (Blocks 1 & 2)
PRIOR_MONTHS = [
    "Hospital Medicine Schedule, 7_1 to 7_31, 2025.html",
    "Hospital Medicine Schedule, 8_1 to 8_31, 2025.html",
    "Hospital Medicine Schedule, 9_1 to 9_30, 2025.html",
    "Hospital Medicine Schedule, 10_1 to 10_31, 2025.html",
    "Hospital Medicine Schedule, 11_1 to 11_30, 2025.html",
    "Hospital Medicine Schedule, 12_1 to 12_31, 2025.html",
    "Hospital Medicine Schedule, 1_1 to 1_31, 2026.html",
    "Hospital Medicine Schedule, 2_1 to 2_28, 2026.html",
]

# Also include partial June 2025 if it exists (end of prior fiscal year)
JUNE_FILE = "Hospital Medicine Schedule, 6_1 to 6_30, 2025.html"


def classify_service(service_name, hours):
    """Classify a service as 'day', 'night', 'swing', or 'exclude'.

    Uses both service name patterns and shift hours to determine type.
    'exclude' services don't count as shifts for anyone.

    Excluded per scheduling team:
    - APP roles (not physician shifts)
    - Night coverage roles (those are APP/resident coverage)
    - Resident/Fellow roles
    - Behavioral medicine
    - Site Director / Admin roles
    - Hospice
    - Coverage roles (Virtua coverage = their version of long call)
    - Kessler Rehab (staffed by APP)
    - Holy Redeemer
    - Cape RMD
    - UM/SAH (utilization management)
    - Long Call (these are extensions, not full shifts)
    - DO NOT USE rows
    - Moonlighting resident roles
    - LTC-SAR roles
    - CCPA
    """
    sname = service_name.lower()
    hours_lower = hours.lower() if hours else ""

    # ── EXCLUDE: Services that don't count as shifts ─────────────────
    # APP roles
    if " app" in sname or sname.startswith("app ") or sname.endswith(" app"):
        return "exclude"
    if "apn" in sname:
        return "exclude"
    if " pa " in sname or sname.endswith(" pa"):
        return "exclude"

    # Night Coverage roles (these are APP/resident coverage, not physician night shifts)
    if "night coverage" in sname:
        return "exclude"
    if "coverage" in sname:
        return "exclude"

    # Resident / Fellow roles
    if "resident" in sname:
        return "exclude"
    if "fellow" in sname:
        return "exclude"
    if "moonlighting" in sname:
        return "exclude"

    # Admin / Leadership roles
    if "admin" in sname:
        return "exclude"
    if "site director" in sname:
        return "exclude"
    if "app lead" in sname:
        return "exclude"

    # Behavioral medicine
    if "behavioral" in sname:
        return "exclude"

    # Hospice
    if "hospice" in sname:
        return "exclude"

    # Kessler Rehab
    if "kessler" in sname:
        return "exclude"

    # Holy Redeemer
    if "holy redeemer" in sname:
        return "exclude"

    # Cape RMD
    if "cape rmd" in sname:
        return "exclude"

    # UM / SAH
    if sname.strip() in ["um", "sah"] or "um rounds" in sname or "um referral" in sname:
        return "exclude"

    # CCPA
    if "ccpa" in sname or "clinical care physician advisor" in sname:
        return "exclude"

    # Long Call roles (extensions, not full shifts)
    if "long call" in sname:
        return "exclude"

    # Teaching Admissions (short shifts, APP-like)
    if "teaching admissions" in sname:
        return "exclude"

    # DO NOT USE
    if "do not use" in sname:
        return "exclude"

    # LTC-SAR
    if "ltc-sar" in sname or "ltc" in sname:
        return "exclude"

    # Cape Extra / CC (admin-like)
    if sname.strip() == "cc":
        return "exclude"

    # Woodbury Consult (APP-like roles)
    if "woodbury" in sname:
        return "exclude"

    # ── Night shifts ────────────────────────────────────────────────────
    # Explicit night keywords in remaining services
    if any(kw in sname for kw in ["night", "nocturnist", "(nah)"]):
        return "night"

    # Hours starting at 5p/7p and ending at 5a/7a = night
    if re.match(r'[57]p-[57]a', hours_lower):
        return "night"
    if re.match(r'5a-7a', hours_lower):
        return "night"
    if re.match(r'11p-7a', hours_lower):
        return "night"

    # ── Swing shifts ────────────────────────────────────────────────────
    if "swing" in sname:
        return "swing"
    # Typical swing hours: 2p-10p, 4p-12a, 3p-12a, 1p-9p
    if re.match(r'[1-4]p-', hours_lower):
        return "swing"

    # ── Day shifts — everything remaining with typical day hours ───────
    return "day"


def normalize_provider_name(name):
    """Clean up provider name from HTML parsing artifacts."""
    if not name:
        return ""
    # Remove footnote numbers at end
    name = re.sub(r'\s*\d+\s*$', '', name).strip()
    # Collapse whitespace
    name = re.sub(r'\s+', ' ', name)
    # Skip non-provider entries
    skip_patterns = ["OPEN SHIFT", "Resident Covering", "Fellow", "DO NOT USE"]
    for sp in skip_patterns:
        if sp.lower() in name.lower():
            return ""
    return name


def main():
    print("=" * 70)
    print("RECALCULATING PRIOR BLOCK ACTUALS FROM HTML SCHEDULES")
    print("Blocks 1 & 2: Jul 2025 through Feb 2026")
    print("=" * 70)

    # Check for June 2025
    prior_files = list(PRIOR_MONTHS)
    june_path = os.path.join(INPUT_DIR, JUNE_FILE)
    if os.path.exists(june_path):
        prior_files.insert(0, JUNE_FILE)
        print(f"Including June 2025 partial month")

    # Parse all prior month HTML files
    all_month_data = []
    for fname in prior_files:
        fpath = os.path.join(INPUT_DIR, fname)
        if not os.path.exists(fpath):
            print(f"WARNING: Missing file: {fname}")
            continue
        print(f"  Parsing: {fname}")
        month_data = parse_schedule(fpath)
        all_month_data.append(month_data)

    # Merge all months
    merged = merge_schedules(all_month_data)
    print(f"\nParsed {len(merged['schedule'])} days, "
          f"{len(merged.get('by_provider', {}))} unique providers")

    # Classify all services
    service_classes = {}
    for s in merged["services"]:
        cls = classify_service(s["name"], s["hours"])
        service_classes[s["name"]] = cls

    # Count day/night/swing classification
    day_svc = sum(1 for c in service_classes.values() if c == "day")
    night_svc = sum(1 for c in service_classes.values() if c == "night")
    swing_svc = sum(1 for c in service_classes.values() if c == "swing")
    print(f"\nService classification: {day_svc} day, {night_svc} night, {swing_svc} swing")

    # Count shifts per provider, DEDUPLICATED per day
    # A provider can appear on multiple service lines per day (umbrella + unit-specific)
    # but that's still ONE shift. We track (provider, date) -> best shift type seen.
    # Priority: night > swing > day (if they have any non-excluded service that day, count once)
    provider_day_shifts = defaultdict(dict)  # provider -> {date_str: shift_type}

    for day_entry in merged["schedule"]:
        dow = day_entry["day_of_week"]
        is_weekend = dow in ("Sat", "Sun")
        date_str = day_entry["date"]

        for assignment in day_entry["assignments"]:
            provider = normalize_provider_name(assignment["provider"])
            if not provider:
                continue

            service_name = assignment["service"]
            svc_class = service_classes.get(service_name, "day")

            if svc_class == "exclude":
                continue

            # Determine shift type for this day
            if is_weekend:
                if svc_class == "night":
                    shift_type = "night"
                elif svc_class == "swing":
                    shift_type = "swing"
                else:
                    shift_type = "weekend"
            else:
                if svc_class == "night":
                    shift_type = "night"
                elif svc_class == "swing":
                    shift_type = "swing"
                else:
                    shift_type = "weekday"

            # Keep the highest-priority shift type for this provider+date
            # Priority: night > swing > weekday/weekend
            existing = provider_day_shifts[provider].get(date_str)
            priority = {"night": 3, "swing": 2, "weekday": 1, "weekend": 1}
            if existing is None or priority.get(shift_type, 0) > priority.get(existing, 0):
                provider_day_shifts[provider][date_str] = shift_type

    # Now aggregate from deduplicated day-level data
    provider_counts = defaultdict(lambda: {
        "weekday_shifts": 0,
        "weekend_shifts": 0,
        "night_shifts": 0,
        "swing_shifts": 0,
    })

    for provider, day_map in provider_day_shifts.items():
        for date_str, shift_type in day_map.items():
            if shift_type == "night":
                provider_counts[provider]["night_shifts"] += 1
            elif shift_type == "swing":
                provider_counts[provider]["swing_shifts"] += 1
            elif shift_type == "weekend":
                provider_counts[provider]["weekend_shifts"] += 1
            else:
                provider_counts[provider]["weekday_shifts"] += 1

    # Calculate weeks/weekends/nights
    # Weeks = weekday_shifts / 5 (decimal, no flooring)
    # Weekends = weekend_shifts / 2 (decimal, no flooring)
    # Nights = raw count (individual nights, NOT divided by 5)
    results = {}
    for provider in sorted(provider_counts.keys()):
        counts = provider_counts[provider]
        results[provider] = {
            "weekday_shifts": counts["weekday_shifts"],
            "prior_weeks": round(counts["weekday_shifts"] / 5, 2),
            "weekend_shifts": counts["weekend_shifts"],
            "prior_weekends": round(counts["weekend_shifts"] / 2, 2),
            "night_shifts": counts["night_shifts"],
            "prior_nights": counts["night_shifts"],  # raw count, not divided
            "swing_shifts": counts["swing_shifts"],
        }

    # Write JSON output
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    json_path = os.path.join(OUTPUT_DIR, "prior_actuals.json")
    with open(json_path, "w") as f:
        json.dump(results, f, indent=2)
    print(f"\nWrote {len(results)} providers to: {json_path}")

    # Print summary
    print(f"\n{'='*70}")
    print(f"{'Provider':<45s} {'WkDay':>6s} {'Wks':>7s} {'WkEnd':>6s} {'WEs':>7s} "
          f"{'Nights':>6s} {'Swing':>6s}")
    print("-" * 90)
    for provider in sorted(results.keys()):
        r = results[provider]
        if r["weekday_shifts"] + r["weekend_shifts"] + r["night_shifts"] + r["swing_shifts"] == 0:
            continue
        print(f"{provider:<45s} {r['weekday_shifts']:>6d} {r['prior_weeks']:>7.2f} "
              f"{r['weekend_shifts']:>6d} {r['prior_weekends']:>7.2f} "
              f"{r['night_shifts']:>6d} "
              f"{r['swing_shifts']:>6d}")

    # ── Update block_schedule_input.xlsx ──────────────────────────────────
    update_xlsx(results)

    return results


def update_xlsx(results):
    """Update block_schedule_input.xlsx with recalculated prior actuals."""
    import openpyxl

    # Manual name aliases: XLSX name -> HTML name
    # These handle spelling variants, reversed names, and nickname differences
    NAME_ALIASES = {
        "CERCEO, ELIZABETH": "CERCEO, LISA",          # Different first name in system
        "DHILLION, JASJIT": "DHILLON, JASJIT",        # Typo: double L
        "DUNN JR, ERNEST CHARLES": "DUNN, E. CHARLES MD",  # Suffix/format difference
        "GORDAN, SABRINA": "GORDON, SABRINA",          # Typo: missing O
        # LEE, GRACE — not found in HTML data (may not have worked prior blocks)
        "OBERDORF, W. ERIC": "OBERDORF, ERIC",         # Middle initial
        "ORATE-DIMAPILIS, CHRISTINA": "DIMAPILIS, CHRISTINA",  # Hyphenated last name
        "RACHOIN, JEAN-SEBASTIEN": "RACHOIN, SEBASTIEN",  # First name variant
        "TROYANOVICH, ESTEBAN": "TROYANOVICH, STEVE",  # Esteban = Steve
        "TUDOR, VLAD": "VLAD, TUDOR",                  # First/last reversed in HTML
        "VIJAYKUMAR, ASHWIN": "VIJAYAKUMAR, ASHVIN",   # Spelling variants
    }

    xlsx_path = os.path.join(OUTPUT_DIR, "block_schedule_input.xlsx")
    if not os.path.exists(xlsx_path):
        print(f"\nWARNING: {xlsx_path} not found — skipping XLSX update")
        return

    print(f"\nUpdating: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["Providers"]

    # Find column indices
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col_name = headers.index("provider_name") + 1
    col_prior_wk = headers.index("prior_weeks_worked") + 1
    col_prior_we = headers.index("prior_weekends_worked") + 1
    col_prior_nt = headers.index("prior_nights_worked") + 1

    print(f"  Columns: name={col_name}, prior_wk={col_prior_wk}, "
          f"prior_we={col_prior_we}, prior_nt={col_prior_nt}")

    # Build name matching index from results
    # HTML names are "Last, First" — sheet names are "LAST, FIRST" (uppercase)
    results_by_upper = {}
    for name, data in results.items():
        results_by_upper[name.upper().strip()] = data
        # Also try without credentials
        clean = re.sub(r'\s+(MD|DO|PA|NP|PA-C|MBBS)\s*$', '', name, flags=re.IGNORECASE).strip()
        results_by_upper[clean.upper()] = data
        # Also try removing trailing hyphen (HTML parsing artifact)
        if name.endswith("-"):
            results_by_upper[name[:-1].upper().strip()] = data
        # Also try removing trailing **
        if name.endswith("**"):
            results_by_upper[name[:-2].upper().strip()] = data

    updated = 0
    not_found = []
    for row in range(2, ws.max_row + 1):
        sheet_name = ws.cell(row, col_name).value
        if not sheet_name:
            continue

        sheet_upper = sheet_name.upper().strip()

        # Check alias first
        if sheet_upper in NAME_ALIASES:
            alias_upper = NAME_ALIASES[sheet_upper]
            matched = results_by_upper.get(alias_upper)
            if matched:
                pass  # use alias match
            else:
                matched = None
        else:
            matched = None

        # Try exact match
        if not matched:
            matched = results_by_upper.get(sheet_upper)

        # Try fuzzy: just last name + first name initial
        if not matched:
            parts = sheet_upper.split(",", 1)
            if len(parts) == 2:
                last = parts[0].strip()
                first = parts[1].strip().split()[0] if parts[1].strip() else ""
                for rname, rdata in results_by_upper.items():
                    rparts = rname.split(",", 1)
                    if len(rparts) == 2:
                        rlast = rparts[0].strip()
                        rfirst = rparts[1].strip().split()[0] if rparts[1].strip() else ""
                        if last == rlast and first and rfirst and first[0] == rfirst[0]:
                            matched = rdata
                            break

        if matched:
            old_wk = ws.cell(row, col_prior_wk).value
            old_we = ws.cell(row, col_prior_we).value
            old_nt = ws.cell(row, col_prior_nt).value

            new_wk = matched["prior_weeks"]
            new_we = matched["prior_weekends"]
            new_nt = matched["prior_nights"]

            # Show changes
            changed = False
            if old_wk != new_wk or old_we != new_we or old_nt != new_nt:
                changed = True

            ws.cell(row, col_prior_wk).value = new_wk
            ws.cell(row, col_prior_we).value = new_we
            ws.cell(row, col_prior_nt).value = new_nt
            updated += 1

            if changed:
                print(f"  {sheet_name:<40s} wk: {old_wk} -> {new_wk}, "
                      f"we: {old_we} -> {new_we}, nt: {old_nt} -> {new_nt}")
        else:
            not_found.append(sheet_name)

    wb.save(xlsx_path)
    print(f"\n  Updated {updated} providers in XLSX")
    if not_found:
        print(f"  {len(not_found)} providers in sheet not matched to HTML data:")
        for nf in not_found:
            print(f"    - {nf}")


if __name__ == "__main__":
    main()
